#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
封样检验Web应用 - V5.8.7
基于 SKILL.md V4.0 (2026-06-23)
实现PDF逐页分析、工程图纸判定规则、产品规格书判定规则
V6.2新增：目录勾选状态检测、料号&物料名称跨表一致性检查
V5.1优化：内存管理（gc.collect）、文本截断、实时状态更新、大文件稳定性提升
V5.3修复：KeyError崩溃防护、Excel错误汇总sheet、大文件稳定性、文件去重
"""

import streamlit as st
import os
import tempfile
import pandas as pd
from datetime import datetime, timedelta
import json
import re
import pdfplumber
import gc  # V5.1: 内存管理 - 显式垃圾回收

# ============================================================
# 标准文件读取
# ============================================================

@st.cache_data(ttl=300)
def load_standards():
    """读取审核标准JSON文件"""
    try:
        with open("inspection_standards.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return None
    except Exception as e:
        return None

# ============================================================
# V5.7 新增：物料编码规则加载
# ============================================================

@st.cache_data(ttl=600)
def load_material_coding_rules():
    """读取物料编码规则JSON文件（由物料编码规则.xlsx生成）"""
    try:
        with open("material_coding_rules.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return None
    except Exception:
        return None

def identify_material_type(part_number, coding_rules):
    """
    V5.7新增：根据料号识别物料类型（结构料/电子料）及小类信息
    参数:
        part_number: 料号字符串（如 K6970000223LA）
        coding_rules: load_material_coding_rules() 返回的字典
    返回:
        dict: {"成功": bool, "类型": "结构料/电子料", "小类": "...", "小小类": "...", "详情": "..."}
    """
    result = {"成功": False, "类型": "未知", "小类": "", "小小类": "", "详情": ""}

    if not part_number or not coding_rules:
        result["详情"] = "料号或编码规则为空"
        return result

    # 提取前5位：字母 + 4位数字（如 K6970）
    m = re.match(r'([A-Za-z])(\d{4})', part_number)
    if not m:
        result["详情"] = f"料号格式无法识别（期望字母+4位数字开头，实际：{part_number[:8]})"
        return result

    prefix = m.group(1).upper()  # K 或 R
    code_4 = prefix + m.group(2)  # 如 K6970

    if prefix not in coding_rules:
        result["详情"] = f"未知前缀：{prefix}（期望 K 或 R）"
        return result

    if code_4 in coding_rules[prefix]:
        info = coding_rules[prefix][code_4]
        result["成功"] = True
        result["类型"] = info.get("类型", "未知")
        result["小类"] = info.get("小类", "")
        result["小小类"] = info.get("小小类", "")
        result["详情"] = f"{result['类型']} / {info.get('小类', '')} / {info.get('小小类', '')}"
    else:
        # 4位代码未找到，返回大类信息
        result["类型"] = "结构料" if prefix == "K" else "电子料"
        result["详情"] = f"{result['类型']}（{code_4} 未在编码规则中找到对应小类，已按前缀判断）"

    return result

# ============================================================
# V4.0 新增：PDF逐页分析引擎
# ============================================================

def analyze_pdf_page_by_page(pdf_path):
    """
    V4.0 步骤2：逐页分析PDF内容类型
    返回：list of dict, 每个dict代表一页的分析结果
    V5.1优化: 限制每页文本存储长度，减少内存占用
    """
    results = []
    all_text = ""
    tables = []  # V5.2: 在同一次PDF打开中同时提取表格
    try:
        # V5.3: pdfplumber内存优化
        # V5.4: pdfplumber 内存优化参数
        _pdf_opts = {}  # 注意：不传额外参数，避免版本兼容问题
        with pdfplumber.open(pdf_path, **_pdf_opts) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                try:  # V5.3: 单页解析异常不影响其他页
                    page_num = page.page_number
                    text = page.extract_text() or ""
                    text_lower = text.lower()
                    all_text += text + "\n"

                    # V5.6: 前3页保留完整文本（封面/目录/物料信息），其他页截断至1500字符
                    if page_num <= 3:
                        text_for_storage = text
                        text_lower_for_storage = text_lower
                    else:
                        text_for_storage = text[:1500] if len(text) > 1500 else text
                        text_lower_for_storage = text_lower[:1500] if len(text_lower) > 1500 else text_lower

                    page_info = {
                        "page_num": page_num,
                        "text": text_for_storage,
                        "text_lower": text_lower_for_storage,
                        "is_cover": False,
                        "is_engineering_drawing": False,
                        "is_bom": False,
                        "is_sample_photo": False,
                        "is_cpk": False,
                        "is_rohs_survey": False,
                        "is_reach_survey": False,
                        "is_product_spec": False,
                        "drawing_number": "",
                        "drawing_version": "",
                        "dimensions": [],
                    }

                    # --- 封面页判定 ---
                    if page_num <= 3:
                        if any(kw in text_lower for kw in ["provisional sample", "临时样品承认书", "sample acknowledgement"]):
                            page_info["is_cover"] = True

                    # --- 工程图纸页判定（V4.0 步骤3规则）---
                    drawing_score = 0
                    # 规则1：含图号（如 NC-XC260522-007）
                    drawing_match = re.search(r'[A-Za-z]{2,}-[A-Za-z0-9\-]{3,}', text)
                    if drawing_match:
                        page_info["drawing_number"] = drawing_match.group(0)
                        drawing_score += 3
                    # 规则2：含版本号
                    version_match = re.search(r'(版本|rev)\s*[：:]?\s*[A-Za-z0-9]', text_lower)
                    if version_match:
                        page_info["drawing_version"] = version_match.group(0)
                        drawing_score += 2
                    # 规则3：含尺寸标注（如 2000±50、5.4±0.2、OD、ID）
                    dim_pattern = r'\d{2,}\s*[±±]\s*\d+|\b(od|id|thk|len)\s*[:：]?\s*\d'
                    if re.search(dim_pattern, text_lower):
                        dims = re.findall(r'\d+\.?\d*\s*[±±]\s*\d+', text)
                        page_info["dimensions"] = dims[:10]
                        drawing_score += 3
                    # 规则4：含标题栏（含 批准/审核/制定）
                    if any(kw in text for kw in ["批准", "审核", "制定", "设计", "approve", "check"]):
                        drawing_score += 1
                    # 规则5：含比例、视角
                    if any(kw in text_lower for kw in ["比例", "scale", "视角", "view", "top view", "side view"]):
                        drawing_score += 1

                    if drawing_score >= 4:
                        page_info["is_engineering_drawing"] = True

                    # --- BOM表页判定 ---
                    if "bom" in text_lower or "物料清单" in text or "bill of material" in text_lower:
                        page_info["is_bom"] = True

                    # --- 样品照片页判定 ---
                    if "附上样品实物" in text or "sample photo" in text_lower or "样品照片" in text:
                        page_info["is_sample_photo"] = True

                    # --- CPK报告页判定（V6.1 增强：支持表格形式/繁体中文/多种格式）---
                    cpk_indicators = [
                        r'cp[kK]\s*[:：=]',                    # cpk: / cpk= / cpk：
                        r'cp[kK]\s*報告',                      # 繁体 CPK報告
                        r'cp[kK]\s*报告',                      # 简体 CPK报告
                        r'cp[kK]\s*report',                   # CPK report
                        r'cp[kK]\s*值',                        # CPK值
                        r'cp[kK]\s*[\(（].*?[\)）]',          # CPK(xxx)
                        r'^.*cp[kK].*$',                       # 任意含CPK的行（宽泛匹配）
                    ]
                    # 宽松匹配：页面文本中只要包含 CPK/Cpk 关键词即判定
                    if re.search(r'cp[kK]', text_lower):
                        # 进一步确认不是偶然出现（至少出现2次或有相关统计关键词）
                        cpk_count = len(re.findall(r'cp[kK]', text_lower))
                        has_cpk_context = any(kw in text_lower for kw in [
                            'usl', 'lsl', 'ppk', 'stddev', 'average', 'range',
                            '规格上限', '规格下限', '标准差', '平均值', '量測', '测量',
                            'dimension', 'tolerance', 'nominal', 'specification'
                        ])
                        if cpk_count >= 2 or has_cpk_context:
                            page_info["is_cpk"] = True

                    # --- RoHS调查表页判定 ---
                    if "rohs" in text_lower and ("调查表" in text or "survey" in text_lower):
                        page_info["is_rohs_survey"] = True

                    # --- REACH调查表页判定 ---
                    if "reach" in text_lower and ("调查表" in text or "survey" in text_lower):
                        page_info["is_reach_survey"] = True

                    # --- 产品规格书/技术规范判定（V4.0 步骤4规则）---
                    spec_score = 0
                    if any(kw in text for kw in ["绝缘电阻", "insulation resistance", "Ω"]):
                        spec_score += 2
                    if any(kw in text for kw in ["PVC", "材质", "material spec", "执行标准", "standard"]):
                        spec_score += 1
                    if any(kw in text_lower for kw in ["100%电气测试", "electrical test", "eia/tia"]):
                        spec_score += 2
                    if spec_score >= 2:
                        page_info["is_product_spec"] = True

                    # V5.4: 只在前30页提取表格 + 仅当页面有关键词时（大幅减少内存）
                    _should_extract_tbl = (
                        page_idx < 30 and (
                            re.search(r'cp[kK]|rohs|reach|catalog|目录|料号|part.?number',
                                      text_lower) or
                            page_info.get("is_cover") or
                            page_info.get("is_cpk") or
                            page_info.get("is_rohs_survey")
                        )
                    )
                    if _should_extract_tbl:
                        try:
                            pt = page.extract_tables()
                            if pt:
                                for t in pt:
                                    # V5.4: 限制单个表格最大行数，防止超大表格撑爆内存
                                    if len(t) <= 200:
                                        tables.append({"page": page.page_number, "table": t})
                        except Exception:
                            pass

                    results.append(page_info)
                except Exception as page_err:
                    # V5.3: 单页解析失败记录错误但继续处理其他页面
                    results.append({
                        "page_num": page_idx + 1 if 'page_idx' in dir() else 0,
                        "text": f"[第{page_idx+1}页解析失败: {str(page_err)[:100]}]",
                        "text_lower": "",
                        "is_cover": False, "is_engineering_drawing": False,
                        "is_bom": False, "is_sample_photo": False,
                        "is_cpk": False, "is_rohs_survey": False,
                        "is_reach_survey": False, "is_product_spec": False,
                        "drawing_number": "", "drawing_version": "",
                        "dimensions": [],
                    })
                    continue

    except Exception as e:
        results.append({"error": str(e), "all_text": all_text, "tables": tables})

    # V5.2: 截断过长的全文文本，避免内存爆炸
    if len(all_text) > 20000:  # V5.4: 进一步降低内存占用
        all_text = all_text[:20000] + "\n[... 文本过长，已截断 ...]"

    return results, all_text, tables  # V5.2: 同时返回预提取的表格


def check_file_type(pdf_path, pdf_text):
    """
    V4.0 第一步：文件类型检查
    返回: (file_type, note)
    file_type: "DQM-002" / "DQM-001" / "unknown"
    """
    text_lower = pdf_text.lower()
    if "xc-r-0802-dqm-002" in text_lower or "正式样品承认书" in text_lower:
        return "DQM-002", "正式样品承认书"
    if "xc-r-0802-dqm-001" in text_lower or "临时样品承认书" in text_lower:
        return "DQM-001", "⚠️ 临时样品承认书（与正式模板不符）"
    # 从文件名判断
    fname = os.path.basename(pdf_path).upper()
    if "DQM-002" in fname:
        return "DQM-002", "从文件名判定为正式样品承认书"
    if "DQM-001" in fname:
        return "DQM-001", "⚠️ 从文件名判定为临时样品承认书"
    return "unknown", "未识别文件类型编号"


def check_engineering_drawing_detailed(page_analysis):
    """
    V4.0 步骤3：工程图纸详细判定
    返回: (is_present, details_dict)
    """
    has_drawing = any(p.get("is_engineering_drawing") for p in page_analysis)
    details = {
        "has_drawing": has_drawing,
        "drawing_pages": [p["page_num"] for p in page_analysis if p.get("is_engineering_drawing")],
        "drawing_numbers": list(set(p.get("drawing_number", "") for p in page_analysis if p.get("drawing_number"))),
        "has_dimensions": any(p.get("dimensions") for p in page_analysis),
        "note": "",
    }
    if has_drawing:
        details["note"] = f"✅ 已提供工程图纸（第 {', '.join(map(str, details['drawing_pages']))} 页）"
        if details["drawing_numbers"]:
            details["note"] += f"，图号：{', '.join(details['drawing_numbers'])}"
    else:
        # 即使未标记为工程图纸，检查是否有含图号的页面
        has_dn = any(p.get("drawing_number") for p in page_analysis)
        if has_dn:
            details["note"] = "⚠️ 页面含图号但未标记为工程图纸，建议人工确认"
            details["has_drawing"] = True  # 保守判定为已提供
        else:
            details["note"] = "❌ 未检测到工程图纸页"
    return details["has_drawing"], details


def check_screw_drawing_requirements(page_analysis, tables=None, material_name=""):
    """
    V5.9.0: 螺丝/紧固件类物料工程图纸特殊要求检查
    检查项：
      1. 螺丝颜色（必填）- 如 BLUE-TY-NY, 黑锌, 白锌, 镍 等
      2. 耐落/防松胶处理（条件必填）- 如有则需标注
      3. 点胶处理（条件必填）- 如有则需标注
    返回: dict with check results
    """
    result = {
        "is_screw": False,
        "checks": [],
        "overall_status": "✅ 通过",
        "issues": [],
    }

    # 判断是否为螺丝类物料
    _screw_keywords = ["螺丝", "螺钉", "螺栓", "screw", "bolt", "fastener", "紧固件"]
    _name_lower = (material_name + " " + " ".join(p.get("text", "")[:200] for p in page_analysis[:3])).lower()
    is_screw = any(kw in _name_lower for kw in _screw_keywords)

    if not is_screw:
        return result

    result["is_screw"] = True

    # 提取图纸页面的文本和表格内容
    all_text = ""
    drawing_texts = []
    for p in page_analysis:
        if p.get("is_engineering_drawing") or p.get("drawing_number"):
            drawing_texts.append(p.get("text", ""))
            all_text += p.get("text", "") + " "
        # 也检查前10页（有些图纸未被正确标记）
        elif p.get("page_num", 0) <= 10:
            all_text += p.get("text", "") + " "

    all_text_lower = all_text.lower()

    # === SF-1: 螺丝颜色（必填）===
    color_check = {"id": "SF-1", "name": "螺丝颜色(Surface Color)", "status": "", "detail": ""}
    _color_keywords = [
        "color", "颜色", "coating", "镀层", "surface", "表面处理",
        "blue", "black", "white", "yellow", "zn", "ni", "cr",
        "ty", "ny", "蓝", "黑", "白", "黄", "镍", "锌",
        "蓝白锌", "白锌", "黄锌", "黑锌", "彩锌"
    ]
    _color_patterns = [
        r'(?:color|颜色|coating|镀层)[\s:\-]*([A-Za-z0-9\-]+)',
        r'(BLUE|BLACK|WHITE|YELLOW|NI|CR)[\s\-]*(TY|NY|ZN)?',
        r'(蓝\s*(白|黑)?\s*锌|白\s*锌|黄\s*锌|黑\s*锌|彩\s*锌|镍)',
        r'([A-Z]{2,}[\-][A-Z]{2,})',  # 如 BLUE-TY-NY
    ]

    color_found = False
    color_value = ""

    # 先在表格中查找（更准确）
    if tables:
        for t_dict in tables:
            tbl_page = t_dict.get("page", 0)
            if tbl_page > 15:
                continue
            tbl = t_dict.get("table", [])
            for row in tbl:
                if not row:
                    continue
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    cell_lower = cell_str.lower()
                    # 匹配颜色关键词行
                    if any(kw in cell_lower for kw in ["color", "颜色", "coating", "镀层"]):
                        # 取同行或附近单元格的值
                        color_found = True
                        color_value = cell_str
                        break
                    # 直接匹配常见颜色格式
                    if re.search(r'(BLUE[-\s]*TY[-\s]*NY|BLACK[-\s]*ZN|WHITE[-\s]*ZN|YELLOW[-\s]*ZN|[A-Z]{2,}[-][A-Z]{2,}|蓝.*锌|白.*锌|黄.*锌|黑.*锌)', cell_str, re.IGNORECASE):
                        color_found = True
                        color_value = cell_str.strip()
                        break
                if color_found:
                    break
            if color_found:
                break

    # 文本兜底搜索
    if not color_found:
        for pat in _color_patterns:
            m = re.search(pat, all_text, re.IGNORECASE)
            if m:
                color_found = True
                color_value = m.group(0).strip()
                break

    if color_found:
        color_check["status"] = "✅ 已注明"
        color_check["detail"] = f"检测到颜色信息：{color_value[:50]}"
    else:
        color_check["status"] = "❌ 缺失"
        color_check["detail"] = "工程图纸未注明螺丝颜色（Surface Color / Coating）"
        result["issues"].append("[SF-1] 工程图纸缺少螺丝颜色标注")

    result["checks"].append(color_check)

    # === SF-2: 耐落处理（条件必填）===
    nylock_check = {"id": "SF-2", "name": "耐落处理(Nylock)", "status": "", "detail": ""}
    _nylock_keywords = ["耐落", "防松", "尼龙圈", "止退胶", "防松胶", "nylock", "nylok", "nylon patch", "anti-loose"]
    has_nylock_ref = any(kw in all_text_lower for kw in _nylock_keywords)

    if has_nylock_ref:
        # 有耐落相关引用，需确认图纸上明确标注
        nylock_check["status"] = "✅ 已注明"
        nylock_check["detail"] = f"检测到耐落处理标注"
    else:
        nylock_check["status"] = "⏭️ 不适用"
        nylock_check["detail"] = "未检测到耐落处理需求，跳过此项"

    result["checks"].append(nylock_check)

    # === SF-3: 点胶处理（条件必填）===
    glue_check = {"id": "SF-3", "name": "点胶处理(Gluing)", "status": "", "detail": ""}
    _glue_keywords = ["点胶", "螺纹胶", "厌氧胶", "loctite", "乐泰", "thread locker", "密封胶", "预涂胶"]
    has_glue_ref = any(kw in all_text_lower for kw in _glue_keywords)

    if has_glue_ref:
        glue_check["status"] = "✅ 已注明"
        glue_check["detail"] = f"检测到点胶处理标注"
    else:
        glue_check["status"] = "⏭️ 不适用"
        glue_check["detail"] = "未检测到点胶处理需求，跳过此项"

    result["checks"].append(glue_check)

    # 整体状态判定
    if color_check["status"] == "❌ 缺失":
        result["overall_status"] = "❌ 不合格（螺丝颜色缺失）"
    else:
        result["overall_status"] = "✅ 通过"

    return result


def check_supplier_info_completeness(page_analysis, tables=None):
    """
    V5.9.2: 封面供应商信息完整性检查
    检查封样承认书封面页的 "Supplier Information（供应商信息）" 区域是否完整填写。
    
    必填字段：
      SI-1 供应商名称（Supplier name）— 必须填写，不能为空或占位符
      SI-2 供应商地址（Supplier address）— 必须填写
      SI-3 联系方式（Supplier contact）— 必须填写真实信息，不能仅为占位符(XXX)
      SI-4 收件人邮箱（Recipient email）— 必须填写
    
    参考字段（不影响合格判定，仅作提示）：
      SI-5 部门签署（Department Signatures）— Produced/Engineering/Design/Quality + 签名+日期
    
    返回: dict {
        "checks": [...],
        "overall_status": "✅ 通过" | "❌ 不合格",
        "issues": [...],
        "supplier_name_found": str or None,
        "cover_page_num": int,
    }
    """
    result = {
        "checks": [],
        "overall_status": "✅ 通过",
        "issues": [],
        "supplier_name_found": None,
        "cover_page_num": 0,
    }

    # --- 定位封面页 ---
    cover_page = None
    cover_text = ""
    for p in page_analysis:
        if p.get("is_cover"):
            cover_page = p
            result["cover_page_num"] = p.get("page_num", 0)
            break

    # 如果未标记is_cover，取前3页作为备选
    if not cover_page:
        for p in page_analysis[:3]:
            txt = p.get("text", "")
            if any(kw in txt.lower() for kw in ["provisional sample", "临时样品承认书", "sample acknowledgement"]):
                cover_page = p
                result["cover_page_num"] = p.get("page_num", 0)
                break

    # 最终兜底：取第1页
    if not cover_page and page_analysis:
        cover_page = page_analysis[0]
        result["cover_page_num"] = cover_page.get("page_num", 1)

    if not cover_page:
        result["overall_status"] = "⏱ 无法检测"
        result["issues"].append("[SI] 无法定位封面页")
        return result

    cover_text = cover_page.get("text", "")
    cover_text_lower = cover_text.lower()

    # === 收集封面表格数据（更准确）===
    cover_tables = []
    if tables:
        cpn = result["cover_page_num"]
        for t_dict in tables:
            tpg = t_dict.get("page", 0)
            # 封面表格通常在page 1~5
            if tpg <= 5:
                tbl = t_dict.get("table", [])
                if tbl:
                    cover_tables.append(tbl)

    # 将所有表格单元格文本合并用于搜索
    all_table_text = ""
    for tbl in cover_tables:
        for row in tbl:
            if row:
                for cell in row:
                    if cell:
                        all_table_text += str(cell) + "\n"

    combined_text = cover_text + "\n" + all_table_text

    # 占位符/无效内容模式
    _placeholder_patterns = [
        r'^xxx[\-\s]*$',           # 纯XXX
        r'^xxx.*$',                 # 以XXX开头
        r'不能填写序列号',          # 提示性文字
        r'placeholder',             # placeholder英文
        r'^\s*[-_]+\s*$',           # 纯横线/下划线
        r'^[Nn][Aa]\s*$',          # N/A
        r'^\s*$',                   # 空白
    ]

    def _is_placeholder(val_str):
        """判断值是否为占位符或无效"""
        v = val_str.strip()
        if not v or len(v) < 2:
            return True
        for pat in _placeholder_patterns:
            if re.match(pat, v, re.IGNORECASE):
                return True
        return False

    def _find_field_value(field_keywords, search_area=None):
        """
        在搜索区域中查找目标字段对应的值
        field_keywords: 字段标签关键词列表，如 ["Supplier name", "供应商名称"]
        返回: (found: bool, value: str, source: str)
        """
        area = search_area or combined_text
        
        # 方法1: 表格查找（最准确）
        for tbl in cover_tables:
            for row_idx, row in enumerate(tbl):
                if not row:
                    continue
                for col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    cell_lower = cell_str.lower()
                    
                    # 找到字段标签行/单元格
                    if any(kw in cell_lower for kw in field_keywords):
                        # 取同行右侧相邻单元格的值
                        for j in range(col_idx + 1, len(row)):
                            if row[j] is not None:
                                val = str(row[j]).strip()
                                if val and not _is_placeholder(val):
                                    return True, val, f"表格(row{row_idx},col{j})"
                        
                        # 取下一行同列的值（字段在上一行，值在下一行）
                        if row_idx + 1 < len(tbl):
                            next_row = tbl[row_idx + 1]
                            if next_row and col_idx < len(next_row) and next_row[col_idx] is not None:
                                val = str(next_row[col_idx]).strip()
                                if val and not _is_placeholder(val):
                                    return True, val, f"表格(row{row_idx+1},col{col_idx})"
        
        # 方法2: 文本正则查找（兜底）
        for kw in field_keywords:
            patterns = [
                rf'{kw}[\s:\-：]*([^\n\r]+)',
                rf'({kw}[^a-zA-Z\u4e00-\u9fff]*)[\s:\-：]*([^\n\r]+)',
            ]
            for pat in patterns:
                m = re.search(pat, area, re.IGNORECASE)
                if m:
                    # 取最后一个捕获组
                    val = (m.group(len(m.groups()))).strip()
                    if val and not _is_placeholder(val):
                        return True, val, "文本正则"
        
        return False, "", ""

    # === SI-1: 供应商名称（必填）===
    si1 = {"id": "SI-1", "name": "供应商名称(Supplier name)", "status": "", "detail": ""}
    found, val, src = _find_field_value(["supplier name", "供应商名称", "供应商名称名称"])
    if found and val:
        si1["status"] = "✅ 已填写"
        si1["detail"] = f"{val[:60]}"
        result["supplier_name_found"] = val[:60]
    else:
        si1["status"] = "❌ 未填写"
        si1["detail"] = "封面「供应商信息」区域缺少供应商名称"
        result["issues"].append("[SI-1] 供应商名称未填写（Supplier name）")
    result["checks"].append(si1)

    # === SI-2: 供应商地址（必填）===
    si2 = {"id": "SI-2", "name": "供应商地址(Supplier address)", "status": "", "detail": ""}
    found, val, src = _find_field_value(["supplier address", "供应商地址", "供应商品质地址"])
    if found and val:
        si2["status"] = "✅ 已填写"
        si2["detail"] = f"{val[:60]}"
    else:
        si2["status"] = "❌ 未填写"
        si2["detail"] = "封面缺少供应商地址信息"
        result["issues"].append("[SI-2] 供应商地址未填写（Supplier address）")
    result["checks"].append(si2)

    # === SI-3: 联系方式/联系人电话（必填，排除占位符）===
    si3 = {"id": "SI-3", "name": "联系方式(Supplier contact)", "status": "", "detail": ""}
    found, val, src = _find_field_value(["supplier contact", "联系方式", "联系人电话", "供应商联系人"])
    if found and val:
        # 额外检查是否为占位符（如纯XXX、手机号占位符等）
        if _is_placeholder(val):
            si3["status"] = "❌ 无效（占位符）"
            si3["detail"] = f"联系方式仅含占位符文字（{val[:30]}），需填写真实联系信息"
            result["issues"].append(f"[SI-3] 联系方式为占位符（{val[:30]}），请填写真实联系方式")
        else:
            si3["status"] = "✅ 已填写"
            si3["detail"] = f"{val[:60]}"
    else:
        si3["status"] = "❌ 未填写"
        si3["detail"] = "封面缺少供应商联系方式"
        result["issues"].append("[SI-3] 联系方式未填写（Supplier contact）")
    result["checks"].append(si3)

    # === SI-4: 收件人邮箱地址（必填）===
    si4 = {"id": "SI-4", "name": "收件人邮箱(Email)", "status": "", "detail": ""}
    found, val, src = _find_field_value(["recipient email", "收件人邮箱", "email address"])
    if found and val:
        # 检查是否像有效email（包含@符号和域名）
        if "@" in val:
            si4["status"] = "✅ 已填写"
            si4["detail"] = f"{val[:50]}"
        elif _is_placeholder(val):
            si4["status"] = "❌ 无效（占位符）"
            si4["detail"] = f"邮箱仅含占位符（{val[:30]}）"
            result["issues"].append(f"[SI-4] 收件人邮箱为占位符（{val[:30]}）")
        else:
            si4["status"] = "✅ 已填写"
            si4["detail"] = f"{val[:50]}"
    else:
        si4["status"] = "❌ 未填写"
        si4["detail"] = "封面缺少收件人邮箱地址"
        result["issues"].append("[SI-4] 收件人邮箱未填写（Email）")
    result["checks"].append(si4)

    # === SI-5: 部门签署（参考项，不判定不合格）===
    si5 = {"id": "SI-5", "name": "部门签署(Department Signatures)", "status": "", "detail": ""}
    sig_keywords = ["produced", "engineering department", "design department", "quality department",
                     "制作人", "工程部", "设计部", "品管部", "签署者", "signatory"]
    date_keywords = ["signing date", "签署日期", "签名日期"]
    
    # 检查是否有签名（手写体通常提取不出，但如果有盖章或打印体名字则可识别）
    sig_found = any(kw in cover_text_lower for kw in sig_keywords)
    date_found = any(kw in cover_text_lower for kw in date_keywords)
    
    # 在表格中查找签名区域是否有实际内容（非空格）
    sig_content_found = False
    for tbl in cover_tables:
        for row in tbl:
            for cell in row:
                if cell:
                    cs = str(cell).strip()
                    # 签名区域可能有日期格式如 2026.6.8 或中文姓名
                    if re.search(r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}', cs) or \
                       (len(cs) >= 2 and len(cs) <= 10 and re.search(r'[\u4e00-\u9fff]{2,}', cs)):
                        sig_content_found = True
                        break
            if sig_content_found:
                break
        if sig_content_found:
            break

    if sig_content_found:
        si5["status"] = "✅ 已签署"
        si5["detail"] = "检测到签名/日期信息"
    elif sig_found and date_found:
        si5["status"] = "⚠️ 有签署区但内容疑似空白"
        si5["detail"] = "存在部门签署区域，建议确认各栏已签字并注明日期"
    else:
        si5["status"] = "⚠️ 未检测到签署信息"
        si5["detail"] = "未检测到部门签署信息（此项为参考项）"
    result["checks"].append(si5)

    # === 整体状态判定 ===
    required_fails = sum(1 for c in result["checks"][:4] if c["status"].startswith("❌"))
    if required_fails >= 3:
        result["overall_status"] = "❌ 不合格（供应商信息严重缺失）"
    elif required_fails >= 1:
        result["overall_status"] = "❌ 不合格（供应商信息不完整）"
    else:
        result["overall_status"] = "✅ 通过"

    return result


def check_product_specification(page_analysis):
    """
    V4.0 步骤4：产品规格书判定
    返回: (status, note)
    status: "独立提供" / "部分提供（嵌入工程图纸）" / "缺失"
    """
    # 检查是否有独立的产品规格书页
    has_spec_page = any(p.get("is_product_spec") for p in page_analysis)
    # 检查工程图纸页是否含技术规格内容
    drawing_pages = [p for p in page_analysis if p.get("is_engineering_drawing") or p.get("drawing_number")]
    has_spec_in_drawing = False
    if drawing_pages:
        for p in drawing_pages:
            txt = p.get("text", "")
            if any(kw in txt for kw in ["绝缘电阻", "PVC", "执行标准", "Ω", "材质", "100%电气"]):
                has_spec_in_drawing = True
                break

    if has_spec_page:
        return "独立提供", "✅ 已提供独立的产品规格书"
    elif has_spec_in_drawing:
        return "部分提供", "⚠️ 技术要求嵌入在工程图纸中，建议补充独立的产品规格书文档"
    else:
        return "缺失", "❌ 未检测到产品规格书内容"


# ============================================================
# PDF解析引擎（基础）
# ============================================================

def extract_all_text(pdf_path):
    """提取PDF全部文本（合并）
    V5.1优化: 限制总文本长度为50000字符，避免超大PDF内存溢出
    """
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
                # V5.1: 超过限制后截断，避免内存爆炸
                if len(text) > 50000:
                    text = text[:50000] + "\n[... 文本过长，已截断 ...]"
                    break
    except Exception as e:
        text = f"[PDF解析错误: {e}]"
    return text


def extract_pdf_tables(pdf_path):
    """提取PDF中的表格"""
    tables = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)
    except Exception:
        pass
    return tables


def extract_dates_from_text(text):
    """从文本中提取所有日期（V5.6增强：支持更多日期格式）"""
    dates = []
    
    # 格式1: YYYY-MM-DD / YYYY.MM.DD / YYYY/MM/DD
    patterns = [
        r'(\d{4})[/\-\.](\d{1,2})[/\-\.](\d{1,2})',
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text)
        for m in matches:
            try:
                date_str = f"{m[0]}-{m[1].zfill(2)}-{m[2].zfill(2)}"
                dates.append(datetime.strptime(date_str, "%Y-%m-%d").date())
            except ValueError:
                continue
    
    # 格式2（V5.6新增）：英文月份格式 "Feb 28, 2026" / "February 28, 2026" / "28-Feb-2026"
    month_names = {
        'jan': 1, 'january': 1, 'feb': 2, 'february': 2,
        'mar': 3, 'march': 3, 'apr': 4, 'april': 4,
        'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
        'aug': 8, 'august': 8, 'sep': 9, 'september': 9,
        'oct': 10, 'october': 10, 'nov': 11, 'november': 11,
        'dec': 12, 'december': 12,
    }
    # "Month DD, YYYY" 或 "Month DD YYYY"
    eng_date_pattern = r'([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})'
    for m in re.findall(eng_date_pattern, text, re.IGNORECASE):
        month_str, day, year = m[0].lower(), int(m[1]), int(m[2])
        if month_str in month_names:
            try:
                dates.append(datetime(year, month_names[month_str], day).date())
            except ValueError:
                continue
    # "DD-Mon-YYYY" 或 "DD/Mon/YYYY"
    short_eng_pattern = r'(\d{1,2})[-/]([A-Za-z]{3})[-/](\d{4})'
    for m in re.findall(short_eng_pattern, text):
        day, month_str, year = int(m[0]), m[1].lower(), int(m[2])
        if month_str in month_names:
            try:
                dates.append(datetime(year, month_names[month_str], day).date())
            except ValueError:
                continue

    # 格式3（V5.6新增）：中文日期 "2026年11月1日" / "2026年06月01日"
    cn_date_pattern = r'(\d{4})年(\d{1,2})月(\d{1,2})日'
    for m in re.findall(cn_date_pattern, text):
        try:
            dates.append(datetime(int(m[0]), int(m[1]), int(m[2])).date())
        except ValueError:
            continue

    return dates


def extract_cpk_values(text, tables=None):
    """
    V6.1 增强：从文本和表格中提取CPK值
    V5.2 优化：接受预提取的 tables，避免重复打开PDF
    支持多种格式：
    1. 文本键值对：cpk: 1.33 / cpk=1.33
    2. 表格形式：CPK列头 + 数值行（使用预提取的tables）
    3. 统计行格式：... | CPK | 0.67 | 3.84 | ...
    4. 宽松相邻匹配：CPK/Cpk 紧邻数字
    """
    cpk_values = []
    seen = set()

    def add_val(v):
        """去重添加CPK值"""
        # V5.8.5: 提高下限从0到0.30，避免误提取公差/偏差等小数值
        if v not in seen and 0.30 <= v <= 100:  # 合理的CPK范围
            seen.add(v)
            cpk_values.append(v)

    # ===== 方法1：文本正则匹配（原有逻辑 + 扩展）=====
    text_patterns = [
        r'cpk\s*[:：=]\s*(\d+\.?\d*)',           # cpk: 1.33
        r'cpk\s*[\(（]\s*(\d+\.?\d*)',           # cpk(1.33)
        r'cpk\s+value\s*[:：=]\s*(\d+\.?\d*)',   # cpk value: 1.33
        r'cp\s*[kK]\s*[:：=]\s*(\d+\.?\d*)',     # CP : 1.33
        r'c pk\s*[:：=]\s*(\d+\.?\d*)',          # c pk: 1.33 (可能的空格)
        # V6.1 新增：宽松相邻匹配
        r'cp[kK]\s*[\/>]?\s*(\d+\.\d{1,2})',     # CPK>1.33 / CPK/1.33 / CPK 1.33
        r'cp[kK]\s+(?:最小|min|minimum)?\s*[:：]?\s*(\d+\.\d+)',  # CPK最小: 1.33
        # V5.8.5修复：此模式在全文搜索，可能跨行误匹配非CPK数字
        # 改为逐行匹配，且要求CPK和数值在同一行内紧密相邻
        # r'(?:cpk|Cpk|CPK)\s.*?(\d+\.\d{2})',  # 已移除：过于宽泛
    ]
    for pattern in text_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for val in matches:
            try:
                add_val(float(val))
            except ValueError:
                continue

    # ===== 方法2：从预提取的表格中提取CPK值（V5.2 优化）=====
    if tables:
        for t_dict in tables:
            table = t_dict["table"]  # pdfplumber extract_tables() 返回的一个表格
            if not table or len(table) < 2:
                continue
            # 寻找包含 CPK 的表头行
            for row_idx, row in enumerate(table):
                if row is None:
                    continue
                row_strs = [str(c).strip() if c else "" for c in row]
                row_text_lower = " ".join(row_strs).lower()
                if re.search(r'cp[kK]', row_text_lower):
                    cpk_col_indices = []
                    for col_idx, cell in enumerate(row_strs):
                        if re.search(r'cp[kK]', cell.lower()):
                            cpk_col_indices.append(col_idx)
                    if cpk_col_indices:
                        for data_row_idx in range(row_idx + 1, min(row_idx + 11, len(table))):
                            data_row = table[data_row_idx]
                            if data_row is None:
                                continue
                            # V5.8.5修复：只提取纯数字单元格，避免从备注/说明文字中误提取嵌入数字
                            for col_idx in cpk_col_indices:
                                cell_val = data_row[col_idx] if col_idx < len(data_row) else None
                                if cell_val is not None:
                                    val_str = str(cell_val).strip()
                                    # 只接受纯数字（不含其他文字）
                                    num_match = re.match(r'^(\d+\.?\d*)$', val_str)
                                    if num_match:
                                        try:
                                            add_val(float(num_match.group(1)))
                                        except ValueError:
                                            pass
                                    # V5.8.5: 移除原来的 elif 分支（会从备注/说明中提取嵌入数字如0.5）
                        break

            # 额外：扫描整个表格中的CPK数值模式
            # V5.8.5: 只从纯数字或明确键值对格式的单元格中提取
            for row in table:
                if row is None:
                    continue
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    # 只匹配明确的 CPK:1.33 格式，不匹配嵌入数字
                    m = re.match(r'^[Cc]p[kK]\s*[:：=]\s*(\d+\.?\d*)$', cell_str, re.IGNORECASE)
                    if m:
                        try:
                            add_val(float(m.group(1)))
                        except ValueError:
                            pass

            # ===== 方法2b（V5.6新增）：Cpk作为行标签，数值在其右侧单元格 =====
            # 格式示例：| Dim | 1 | 2 | 3 | 4 | 5 |
            #           | ... | ... | ... | ... | ... | ... |
            #           | Cpk | 1.66 | 1.56 | 1.69 | 1.73 | 1.56 |
            for row_idx, row in enumerate(table):
                if row is None:
                    continue
                row_strs = [str(c).strip() if c else "" for c in row]
                # 寻找Cpk所在的单元格位置
                cpk_cell_idx = None
                for col_idx, cell in enumerate(row_strs):
                    if re.match(r'^[Cc]p[kK]$', cell.strip()):
                        cpk_cell_idx = col_idx
                        break
                if cpk_cell_idx is not None:
                    # 读取该行中Cpk右侧所有单元格的数值
                    for col_idx in range(cpk_cell_idx + 1, len(row_strs)):
                        val = row_strs[col_idx].strip()
                        if not val:
                            continue
                        # V5.8.5: 只接受纯数字单元格
                        num_match = re.match(r'^(\d+\.?\d*)$', val)
                        if num_match:
                            try:
                                v = float(num_match.group(1))
                                if 0.30 <= v <= 100:
                                    add_val(v)
                            except ValueError:
                                pass
                        # V5.8.5: 移除原来的 elif 嵌入数字提取分支

    # ===== 方法3：逐行统计摘要匹配（V5.8.5修复：改为逐行匹配，避免跨行误提取）=====
    # 匹配类似 "CPK  0.67  3.84  4.00" 或 "Cpk 1.36 1.49 2.44 2.93 1.91"
    # V5.8.5 关键修复：原逻辑在全文上做正则，会把非CPK行的数字（如公差0.10）误当CPK值
    _cpk_line_pattern = r'^.*\bcp[kK]\b[\s\:：]*([0-9.]+\s*(?:[0-9.]+\s*)*)'
    for text_line in text.split('\n'):
        text_line_stripped = text_line.strip()
        _line_match = re.search(_cpk_line_pattern, text_line_stripped, re.IGNORECASE)
        if _line_match:
            _match_group = _line_match.group(1)
            # 只从该行提取数字
            numbers = re.findall(r'(\d+\.\d{1,2})', _match_group)
            for n in numbers:
                try:
                    v = float(n)
                    # V5.8.5: 提高最小阈值（CPK值通常>=0.30；<0.3的多为噪声如公差、偏差等）
                    if 0.30 <= v <= 20:
                        add_val(v)
                except ValueError:
                    continue

    return cpk_values


def check_keyword_in_text(text, keywords):
    """检查文本中是否包含关键词列表中的任一关键词"""
    for kw in keywords:
        if kw.lower() in text.lower():
            return True, kw
    return False, None


# ============================================================
# 审核核心逻辑
# ============================================================

def determine_material_type(file_name, standards, pdf_content_hint=None):
    """
    第零步：判定物料类型（电子料 vs 结构件）
    V5.5增强: 支持基于文件名和PDF内容双重判定
    返回: (type_str, type_cn, matched_keyword, needs_electrical_test)
    """
    file_lower = file_name.lower()
    # V5.5: 合并文件名和PDF内容用于关键词匹配
    search_text = file_lower
    if pdf_content_hint:
        search_text = file_lower + " " + pdf_content_hint.lower()

    electronic_kw = standards.get("material_types", {}).get("electronic", {}).get("keywords", [])
    structural_kw = standards.get("material_types", {}).get("structural", {}).get("keywords", [])

    # 先检查电子料关键词（优先在合并文本中搜索）
    for kw in electronic_kw:
        if kw.lower() in search_text:
            required_electrical = standards.get("material_types", {}).get(
                "electronic", {}
            ).get("required_electrical_test", [])
            needs_elec = any(e_kw.lower() in search_text for e_kw in required_electrical)
            source = "(PDF内容)" if pdf_content_hint and kw.lower() not in file_lower and kw.lower() in pdf_content_hint.lower() else "(文件名)"
            return "electronic", "电子料", kw, needs_elec

    # 再检查结构件关键词
    for kw in structural_kw:
        if kw.lower() in search_text:
            return "structural", "结构件", kw, False

    return "unknown", "未知类型", None, False


def inspect_file_completeness_v4(page_analysis, material_type, standards):
    """
    V4.0 第一类：文件完整性检验（基于逐页分析结果）
    返回: dict with item-level results and overall status
    """
    results = {
        "items": [],
        "pass_count": 0,
        "fail_count": 0,
        "total": 0,
        "status": "⚠️ 部分通过",
        "page_analysis_summary": {},
    }

    type_key = "electronic" if material_type == "electronic" else "structural"
    items = standards.get("file_completeness", {}).get(type_key, {}).get("items", [])

    results["total"] = len(items)

    # 基于逐页分析结果判断
    has_drawing = any(p.get("is_engineering_drawing") or p.get("drawing_number") for p in page_analysis)
    has_sample_photo = any(p.get("is_sample_photo") for p in page_analysis)
    has_cpk = any(p.get("is_cpk") for p in page_analysis)
    has_bom = any(p.get("is_bom") for p in page_analysis)
    has_rohs = any(p.get("is_rohs_survey") for p in page_analysis)
    has_reach = any(p.get("is_reach_survey") for p in page_analysis)
    spec_status, spec_note = check_product_specification(page_analysis)

    # 全文本用于其他检查
    all_text = " ".join(p.get("text", "") for p in page_analysis).lower()

    for item in items:
        name = item["name"]
        english = item.get("english", "")
        note = item.get("note", "")
        required = item.get("required", True)

        found = False
        item_note = note or ""

        # 基于页面分析结果判断
        if "物料清单" in name or "Bill of Material" in english:
            found = has_bom
        elif "工程图纸" in name or "Engineering Drawing" in english:
            found, draw_details = check_engineering_drawing_detailed(page_analysis)
            item_note = draw_details.get("note", item_note)
        elif "样品照片" in name or "Sample Photos" in english:
            found = has_sample_photo
        elif "全尺寸" in name or "Full Size" in english:
            # 全尺寸报告通常含尺寸测量表格
            found = "全尺寸" in all_text or "full size" in all_text or "measurement report" in all_text
        elif "Cpk" in name or "Cpk Report" in english:
            found = has_cpk
        elif "产品规格书" in name or "Product Specification" in english:
            found = spec_status != "缺失"
            item_note = spec_note
        elif "制造流程图" in name or "Process Flow" in english:
            found = "制造流程" in all_text or "process flow" in all_text
        elif "包装方式" in name or "Packaging" in english:
            found = "包装" in all_text or "packaging" in all_text
        elif "QC工程图" in name or "QC Flow" in english:
            found = "qc flow" in all_text or "qc 工程" in all_text
        elif "电气性能" in name or "Electrical Performance" in english:
            found = "电气性能" in all_text or "electrical performance" in all_text
        elif "可靠性" in name or "Reliability" in english:
            found = "可靠性" in all_text or "reliability" in all_text
        elif "材质证明" in name or "Material Certificate" in english:
            found = "材质证明" in all_text or "material certificate" in all_text or "sgs" in all_text
        elif "RoHS" in name and "调查表" in name:
            found = has_rohs
        elif "RoHS" in name and "测试报告" in name:
            found = "rohs" in all_text and "test report" in all_text
        elif "REACH" in name and "调查表" in name:
            found = has_reach
        elif "REACH" in name and "测试报告" in name:
            found = "reach" in all_text and "test report" in all_text
        else:
            # 通用关键词匹配
            fnd, _ = check_keyword_in_text(all_text, [name, english])
            found = fnd

        item_result = {
            "序号": item["id"],
            "项目": f"{name} ({english})",
            "必填": "✅" if required is True else ("⚠️" if required == "conditional" else "❌"),
            "结果": "✅ 已找到" if found else "❌ 缺失",
            "备注": item_note,
        }

        if found:
            results["pass_count"] += 1
        else:
            results["fail_count"] += 1

        results["items"].append(item_result)

    if results["fail_count"] == 0:
        results["status"] = "✅ 通过"
    elif results["fail_count"] > 3:
        results["status"] = "❌ 不合格"

    return results


def inspect_rohs_compliance(page_analysis, standards, check_date, tables=None):
    """
    第二类：RoHS合规性检验（4子项）
    V5.8.7修复: 新增tables参数，支持从预提取表格中搜索红框字段
    """
    results = {
        "sub_items": {},
        "overall_status": "⚠️ 部分通过",
        "issues": [],
    }

    all_text = " ".join(p.get("text", "") for p in page_analysis)
    all_text_lower = all_text.lower()

    # 2.1 RoHS 2.0测试报告是否存在（V5.6增强：扩展关键词+表格检测）
    rohs_keywords = [
        "rohs 2.0 test report", "rohs测试报告", "rohs 2.0 report",
        "rohs 2.0 restricted substances", "rohs 2.0限制物质",
        "rohs 2.0 test report",  # 大小写变体
        "restricted substances composition",  # 英文调查表标题
        "sgs", "test report",  # SGS测试报告标识
        # V5.7新增：文件名形式和更多变体
        "rohs.pdf", "rohs .pdf", "限用物质成分调查表",
        "rohs 2.0 questionnaire", "rohs composition",
        "material certificate", "材质证明", "sgs报告",
    ]
    rohs_report_found = any(kw in all_text_lower for kw in rohs_keywords)

    # V5.6额外检查：是否有页面被标记为RoHS页
    if not rohs_report_found:
        rohs_report_found = any(p.get("is_rohs_survey") for p in page_analysis)
    results["sub_items"]["2.1_RoHS测试报告"] = (
        "✅ 已提供" if rohs_report_found else "❌ 未提供"
    )
    if not rohs_report_found:
        results["issues"].append("缺少RoHS 2.0测试报告")

    # 2.2 & 2.4 报告日期有效性
    all_dates = extract_dates_from_text(all_text)
    survey_date_valid = False
    test_date_valid = False

    if all_dates:
        latest_date = max(all_dates)
        days_ago = (check_date.date() - latest_date).days
        if days_ago <= 365:
            survey_date_valid = True
            test_date_valid = True
        else:
            results["issues"].append(
                f"RoHS报告日期过期：报告日期 {latest_date}，距今 {days_ago} 天（超过365天限制）"
            )

    results["sub_items"]["2.2_RoHS调查表日期有效性"] = (
        f"✅ 有效（≤365天）" if survey_date_valid else ("❌ 过期" if all_dates else "⏱️ 未检测到日期")
    )
    results["sub_items"]["2.4_RoHS测试报告日期有效性"] = (
        f"✅ 有效（≤365天）" if test_date_valid else ("❌ 过期" if all_dates else "⏱ 未检测到日期")
    )

    # 2.3 RoHS调查表红框字段（6个必填字段）
    # 检查页面中含RoHS调查表的页面
    rohs_survey_pages = [p for p in page_analysis if p.get("is_rohs_survey")]
    red_box_fields = [
        ("Monomer（单体）", ["monomer", "单体"]),
        ("Supplier（供应商）", ["supplier", "供应商"]),
        ("Control method（控制方法）", ["control method", "控制方法"]),
        ("Number（编号）", ["number", "编号", "report number"]),
        ("Effective date of report（出报告日期）", ["effective date", "出报告日期", "报告日期"]),
        ("Remarks（备注）", ["remarks", "备注"]),
    ]
    missing_fields = []
    for field_name, keywords in red_box_fields:
        found = False
        # V5.8.5：先在RoHS调查表页文本中查找
        for p in rohs_survey_pages:
            txt = p.get("text", "")
            txt_lower = txt.lower()
            # V5.7：使用更宽松的子串匹配
            if any(kw.lower() in txt_lower for kw in keywords):
                found = True
                break
            # V5.8.5增强：多词关键字拆分匹配（如 "effective date" → "effective" 和 "date" 都需出现）
            for kw in keywords:
                kw_words = kw.lower().split()
                if len(kw_words) >= 2:
                    all_words_found = all(w in txt_lower for w in kw_words if len(w) > 1)
                    if all_words_found:
                        found = True
                        break
                else:
                    if kw.lower() in txt_lower:
                        found = True
                        break
            if found:
                break

        # V5.8.5：如果页面文本未找到，在RoHS表格中直接搜索（表格提取更准确）
        if not found and tables:
            for t_dict in tables:
                tbl = t_dict.get("table", [])
                tbl_page = t_dict.get("page", 0)
                if not tbl:
                    continue
                # 只检查RoHS相关表格（前30页内）
                tbl_preview = ""
                for row in tbl[:2]:
                    if row:
                        tbl_preview += " ".join(str(c) if c else "" for c in row) + " "
                if "rohs" not in tbl_preview.lower():
                    continue
                # 在所有单元格中搜索关键词
                for row in tbl:
                    if not row:
                        continue
                    for cell in row:
                        if cell is None:
                            continue
                        cell_str = str(cell).strip().lower()
                        if any(kw.lower() in cell_str for kw in keywords):
                            found = True
                            break
                        # 多词拆分匹配
                        for kw in keywords:
                            kw_words = kw.lower().split()
                            if len(kw_words) >= 2:
                                if all(w in cell_str for w in kw_words if len(w) > 1):
                                    found = True
                                    break
                    if found:
                        break
                if found:
                    break

        # V5.8.5：最后在全文中兜底搜索
        if not found:
            for kw in keywords:
                kw_words = kw.lower().split()
                if len(kw_words) >= 2:
                    all_words_found = all(w in all_text_lower for w in kw_words if len(w) > 1)
                    if all_words_found:
                        found = True
                        break
                else:
                    if kw.lower() in all_text_lower:
                        found = True
                        break
            if not found:
                missing_fields.append(field_name.split("（")[0])

    fill_status = f"✅ 全部填写（{6 - len(missing_fields)}/6）" if len(missing_fields) == 0 else f"❌ {len(missing_fields)}项未填写：{'、'.join(missing_fields)}"
    results["sub_items"]["2.3_RoHS调查表红框字段"] = fill_status
    if missing_fields:
        results["issues"].append(f"RoHS调查表红框字段未填写：{'、'.join(missing_fields)}")

    passed_all = rohs_report_found and survey_date_valid and test_date_valid and len(missing_fields) == 0
    results["overall_status"] = "✅ 通过" if passed_all else "❌ 不合格"

    return results


def inspect_cpk_compliance(page_analysis, standards, pdf_text, tables=None):
    """
    第三类：CPK合规性检验（2子项）
    V6.1: 支持从PDF表格中提取CPK值
    V5.2: 接受预提取的表格，避免重复打开PDF
    """
    results = {
        "sub_items": {},
        "overall_status": "⚠️ 部分通过",
        "issues": [],
        "cpk_values": [],
    }

    all_text = " ".join(p.get("text", "") for p in page_analysis)

    # 3.1 提取CPK值（V6.1: 传入tables以避免重复打开PDF）
    cpk_values = extract_cpk_values(all_text, tables=tables)
    results["cpk_values"] = cpk_values

    if cpk_values:
        min_val = min(cpk_values)
        max_val = max(cpk_values)
        all_pass = all(v >= 1.33 for v in cpk_values)
        results["sub_items"]["3.1_CPK值"] = (
            f"✅ 通过（范围: {min_val:.2f} ~ {max_val:.2f}, 均≥1.33）"
            if all_pass
            else f"❌ 不合格（最小值 {min_val:.2f} < 1.33）"
        )
        if not all_pass:
            results["issues"].append(f"CPK值不合格：最小值 {min_val:.2f}，要求 ≥ 1.33")
    else:
        results["sub_items"]["3.1_CPK值"] = "⏱ 未检测到CPK数据"
        results["issues"].append("未在PDF中找到CPK值数据")

    # 3.2 CPK报告与图纸/全尺寸报告尺寸对应性
    has_cpk_page = any(p.get("is_cpk") for p in page_analysis)
    has_drawing = any(p.get("is_engineering_drawing") or p.get("drawing_number") for p in page_analysis)
    results["sub_items"]["3.2_CPK尺寸对应性"] = (
        "✅ 已检测到CPK及工程图纸页面（详细对应性需人工确认）"
        if has_cpk_page and has_drawing
        else "⏱ 需人工确认CPK与图纸对应关系"
    )

    has_critical_issue = any("CPK值不合格" in issue or "未找到CPK" in issue for issue in results["issues"])
    results["overall_status"] = "✅ 通过" if not has_critical_issue else "❌ 不合格"

    return results


def inspect_dimension_correspondence(page_analysis, standards):
    """
    第四类：尺寸公差对应性检验（包含关系 C ⊆ B ⊆ A）
    """
    results = {
        "sub_items": {},
        "overall_status": "⏱ 待人工确认",
        "issues": [],
    }

    has_drawing = any(p.get("is_engineering_drawing") or p.get("drawing_number") for p in page_analysis)
    has_fullsize = any(
        "全尺寸" in p.get("text", "") or "full size" in p.get("text", "").lower() or "measurement" in p.get("text", "").lower()
        for p in page_analysis
    )
    has_cpk = any(p.get("is_cpk") for p in page_analysis)

    results["sub_items"]["4.1_规格图纸(A层)"] = (
        "✅ 已提供" if has_drawing else "❌ 缺失"
    )
    results["sub_items"]["4.2_全尺寸量测报告(B层)"] = (
        "✅ 已提供" if has_fullsize else "❌ 缺失"
    )
    results["sub_items"]["4.3_CPK报告(C层)"] = (
        "✅ 已提供" if has_cpk else "❌ 缺失"
    )

    all_present = has_drawing and has_fullsize and has_cpk
    if all_present:
        results["sub_items"]["4.4_包含关系C⊆B⊆A"] = (
            "⏱ 三文件均存在，建议人工核对具体尺寸及公差是否一致"
        )
        results["overall_status"] = "⚠️ 需人工确认"
        results["issues"].append("三文件均存在，需人工核对尺寸及公差对应关系")
    else:
        missing = []
        if not has_drawing:
            missing.append("规格图纸")
        if not has_fullsize:
            missing.append("全尺寸量测报告")
        if not has_cpk:
            missing.append("CPK报告")
        results["sub_items"]["4.4_包含关系C⊆B⊆A"] = f"❌ 文件缺失，无法验证包含关系"
        results["overall_status"] = "❌ 不合格"
        results["issues"].append(f"缺失文件：{'、'.join(missing)}，无法验证尺寸包含关系")

    return results


def inspect_report_validity(page_analysis, standards, check_date):
    """
    第五类：报告时效性检验
    """
    results = {
        "sub_items": {},
        "overall_status": "✅ 通过",
        "issues": [],
    }

    all_text = " ".join(p.get("text", "") for p in page_analysis)
    all_dates = extract_dates_from_text(all_text)

    if all_dates:
        latest_date = max(all_dates)
        days_ago = (check_date.date() - latest_date).days

        if days_ago <= 365:
            results["sub_items"]["5.1_RoHS报告时效性"] = (
                f"✅ 有效（报告日期: {latest_date}，距今 {days_ago} 天 ≤ 365 天）"
            )
        else:
            results["sub_items"]["5.1_RoHS报告时效性"] = (
                f"❌ 已过期（报告日期: {latest_date}，距今 {days_ago} 天 > 365 天）"
            )
            results["issues"].append(f"RoHS报告已过期 {days_ago - 365} 天")
            results["overall_status"] = "❌ 不合格"

        # REACH报告
        has_reach = any(p.get("is_reach_survey") for p in page_analysis)
        if has_reach:
            results["sub_items"]["5.2_REACH报告时效性"] = (
                f"✅ 有效（同上）" if days_ago <= 365 else f"❌ 已过期"
            )
        else:
            results["sub_items"]["5.2_REACH报告时效性"] = "⏱ 未检测到REACH报告"
    else:
        results["sub_items"]["5.1_RoHS报告时效性"] = "⏱ 未检测到报告日期"
        results["sub_items"]["5.2_REACH报告时效性"] = "⏱ 未检测到报告日期"

    return results


# ============================================================
# V6.2 新增：目录勾选状态检测 + 料号&物料名称跨表一致性检查
# ============================================================

def extract_cover_info(page_analysis, pdf_path, tables=None):
    """
    V6.2: 从封面/样品承认书页提取料号和物料名称
    V5.7增强：当封面页为图片(无文字)时，从前几页表格中回退搜索料号
    返回: {"part_number": str, "material_name": str, "page_num": int}
    """
    result = {"part_number": "", "material_name": "", "page_num": 0}

    # 寻找封面页或样品承认书页
    cover_pages = [p for p in page_analysis if p.get("is_cover")]
    if not cover_pages:
        # 尝试从含物料信息的页面提取
        cover_pages = [p for p in page_analysis if any(
            kw in p.get("text", "") for kw in [
                "material number", "物料编号", "product name", "产品名称",
                "sample acknowledgement", "样品承认书", "part number", "料号"
            ]
        )]

    # V5.8.5 新增：优先从封面表格中提取（表格格式封面比文本提取更准确）
    _cover_page_nums = set(p.get("page_num", 0) for p in cover_pages)
    if tables and (not result["part_number"] or not result["material_name"]):
        for t_dict in tables:
            tbl_page = t_dict.get("page", 0)
            if tbl_page > 5:  # 只搜索前5页的表格
                continue
            tbl = t_dict.get("table", [])
            if not tbl or len(tbl) < 2:
                continue
            # 检查是否为样品承认书/封面类表格
            _tbl_preview = ""
            for row in tbl[:3]:
                if row:
                    _tbl_preview += " ".join(str(c) if c else "" for c in row) + " "
            _is_ack_tbl = any(kw in _tbl_preview.lower() for kw in [
                "sample acknowledgement", "样品承认书", "product name",
                "material number", "产品名称", "物料编号", "supplier product model"
            ])
            if not _is_ack_tbl:
                continue
            # 逐行扫描表格，找 "Product name" / "产品名称" 标签行，取同行下一个非空单元格作为值
            for row_idx, row in enumerate(tbl):
                if not row:
                    continue
                row_str = " ".join(str(c) if c else "" for c in row)
                # 查找料号：Material number 行
                for ci, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    # 料号匹配：K+数字 格式 或 Material number 标签行
                    if not result["part_number"]:
                        pn_in_cell = re.search(r'(K\d{6,}[A-Za-z]*)', cell_str)
                        if pn_in_cell:
                            result["part_number"] = pn_in_cell.group(1).strip()
                            result["page_num"] = tbl_page
                        elif re.search(r'material\s*(number|no\.?)|物料编号|part\s*number', cell_str, re.IGNORECASE):
                            # 标签单元格，值在右侧相邻单元格
                            for nc in row[ci+1:]:
                                if nc is not None and str(nc).strip():
                                    nv = re.search(r'(K\d{6,}[A-Za-z]*)', str(nc).strip())
                                    if nv:
                                        result["part_number"] = nv.group(1).strip()
                                        result["page_num"] = tbl_page
                                    break

                    # 物料名称匹配：Product name / 产品名称 标签行
                    if not result["material_name"]:
                        if re.search(r'product\s*name|产品名称', cell_str, re.IGNORECASE):
                            # 标签单元格，值在右侧相邻单元格
                            for nc in row[ci+1:]:
                                if nc is not None and str(nc).strip():
                                    name_candidate = str(nc).strip()
                                    # 过滤：不能是标签文字、长度合理
                                    _bad_kw = ['supplier', 'model', 'number', 'remark',
                                               'version', 'description', '规格', '日期',
                                               'rev', '料号', '物料', '编号']
                                    _nc_lower = name_candidate.lower()
                                    if (len(name_candidate) >= 2 and len(name_candidate) <= 60
                                            and not any(k in _nc_lower for k in _bad_kw)
                                            and not name_candidate.isdigit()):
                                        result["material_name"] = name_candidate
                                        result["page_num"] = tbl_page
                                    break
                        elif re.search(r'supplier\s*product\s*model|供應商產品型號|供应商产品型号|型号', cell_str, re.IGNORECASE):
                            for nc in row[ci+1:]:
                                if nc is not None and str(nc).strip():
                                    name_candidate = str(nc).strip()
                                    _bad_kw = ['supplier', 'model', 'number', 'remark', 'version']
                                    if (len(name_candidate) >= 2 and len(name_candidate) <= 50
                                            and not any(k in name_candidate.lower() for k in _bad_kw)
                                            and not name_candidate.isdigit()):
                                        result["material_name"] = name_candidate
                                        result["page_num"] = tbl_page
                                    break

                if result["part_number"] and result["material_name"]:
                    break
            if result["part_number"] and result["material_name"]:
                break

    # 传统方式：从文字行提取（当表格提取未成功时）
    if not result["part_number"] or not result["material_name"]:
        for p in cover_pages:
            text = p.get("text", "")
            if not text:
                continue

            lines = text.split('\n')
            for i, line in enumerate(lines):
                line_stripped = line.strip()

                # 提取料号（多种格式，V5.6增强）
                # 注意：patterns仅用于文档说明，实际匹配在下面单独进行

                # 料号匹配（V5.6增强：支持冒号/等号/括号等多种分隔符）
                pn_match = re.search(
                    r'(?:Material\s*(?:number|No\.?|编号)?|物料编号|Part\s*Number[\(（]?料号[\)）]?|料号[\/\s]*:?|零件号)[\s:：()（）\-*]*([A-Za-z0-9][A-Za-z0-9_\-]*)',
                    line, re.IGNORECASE
                )
                # V5.6额外尝试：直接匹配 K+数字 字母混合格式（常见料号格式）
                if not pn_match:
                    pn_match = re.search(
                        r'\b(K\d+[A-Za-z]*)\b',  # 如 K6970000223LA
                        line
                    )
                if pn_match and not result["part_number"]:
                    result["part_number"] = pn_match.group(1).strip()
                    result["page_num"] = p["page_num"]

                # 物料名称匹配（V5.8.3修复：优先Product name，回退Supplier Product Model）
                _label_keywords = ['supplier', 'model', 'number', 'remark', 'material',
                                   'version', 'description', '规格', '日期', 'date',
                                   'rev', '料号', '物料', '编号', '版本']
                name_match = re.search(
                    r'(?:Product\s*(?:name|名称)|产品名称|Description|零件名称|物料名称|Part\s*name\s*/?\s*model|零件名称/型号)[\s:：\s]*(.+?)\s{2,}',
                    line, re.IGNORECASE
                )
                if not name_match:
                    name_match = re.search(
                        r'(?:Product\s*(?:name|名称)|产品名称|Description|零件名称|物料名称|Part\s*name\s*/?\s*model|零件名称/型号)[\s:：\s]*([^\n\r]{3,60})',
                        line, re.IGNORECASE
                    )
                if name_match and not result["material_name"]:
                    name_val = name_match.group(1).strip()
                    name_lower = name_val.lower()
                    _is_label_like = any(kw in name_lower for kw in _label_keywords)
                    if (len(name_val) >= 3 and not name_val.isdigit()
                            and not _is_label_like
                            and len(name_val) <= 60):
                        result["material_name"] = name_val

            if result["part_number"] or result["material_name"]:
                break

        # V5.8.3增强：如果Product name未提取到，尝试从Supplier Product Model字段提取
        if not result["material_name"]:
            for p in cover_pages:
                text = p.get("text", "")
                if not text:
                    continue
                lines = text.split('\n')
                for line in lines:
                    spm_match = re.search(
                        r'(?:Supplier\s*Product\s*Model|供應商產品型號?|供应商产品型号?|型号)[^:\w]*[\s:：]*([A-Za-z0-9_\-&+\u4e00-\u9fff]{2,50})',
                        line, re.IGNORECASE
                    )
                    if spm_match:
                        candidate = spm_match.group(1).strip()
                        _spm_label_kw = ['supplier', 'model', 'number', 'remark', 'version']
                        if (len(candidate) >= 2 and len(candidate) <= 50
                                and not any(k in candidate.lower() for k in _spm_label_kw)
                                and not candidate.isdigit()):
                            result["material_name"] = candidate
                            break
                if result["material_name"]:
                    break

    # V5.7增强：封面页为图片(无文字)或提取失败时，从前10页表格中搜索料号和物料名称
    if (not result["part_number"] or not result["material_name"]) and tables:
        import re as _re
        for t_dict in tables:
            tbl_page = t_dict.get("page", 0)
            if tbl_page > 10:  # 只搜索前10页
                continue
            tbl = t_dict.get("table", [])
            if not tbl:
                continue
            # 检查是否为封面/样品承认书表格（优先从这类表格提取）
            _tbl_text = ""
            for row in tbl[:3]:
                if row:
                    _tbl_text += " ".join(str(c) if c else "" for c in row) + " "
            _is_cover_tbl = any(kw in _tbl_text.lower() for kw in [
                "sample acknowledgement", "样品承认书", "product name",
                "material number", "产品名称", "物料编号"
            ])

            # 在表格中搜索 K+数字 字母格式（如 K6970000223LA）
            for row in tbl:
                if not row:
                    continue
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    # 提取料号
                    pn_match = _re.search(r'(K\d{6,}[A-Za-z]*)', cell_str)
                    if pn_match and not result["part_number"]:
                        result["part_number"] = pn_match.group(1).strip()
                        result["page_num"] = tbl_page
                    # V5.8.2增强：提取物料名称（支持中英文名称）
                    if not result["material_name"] and _is_cover_tbl:
                        # 匹配常见物料名称格式（如 S1652_FPC连接器泡棉）
                        _name_patterns = [
                            r'^([A-Za-z0-9_\-&+\u4e00-\u9fff]{4,40})$',  # 独立的名称单元格
                            r'(?:Product\s*name|产品名称|Part\s*Name|零件名称)[^:\w]*[\s:：]*([\w\u4e00-\u9fff\-_&+]{2,40})',
                            r'(?:Supplier\s*Product\s*Model|供应商品号型号|型号)[^:\w]*[\s:：]*([\w\u4e00-\u9fff\-_&+]{2,40})',
                        ]
                        for _np in _name_patterns:
                            _nm = _re.search(_np, cell_str, _re.IGNORECASE)
                            if _nm:
                                _candidate = _nm.group(1).strip()
                                # 过滤掉标签类文字
                                _label_kw = ['supplier', 'model', 'number', 'remark',
                                            'material', 'version', 'description']
                                if (len(_candidate) >= 2 and len(_candidate) <= 50
                                        and not any(k in _candidate.lower() for k in _label_kw)
                                        and not _candidate.isdigit()):
                                    result["material_name"] = _candidate
                                    break

                if result["part_number"] and result["material_name"]:
                    break
            if result["part_number"] and result["material_name"]:
                break

    return result


def extract_table_headers_part_info(page_analysis, tables=None):
    """
    V5.8.7: 从各表头提取料号和物料名称（大幅改进）
    改进点：
    1. 跳过表头标签行（含多个label关键词的行）
    2. 识别并跳过非目标字段（供应商名/版本号/项目名/供应商料号等）
    3. 处理跨单元格料号拆分（pdfplumber可能将长料号拆为多个单元格）
    4. 物料名称必须包含中文或下划线格式（排除纯英文短值如S0521/V1.0）
    V5.4优化: 接受预提取的tables，避免重新打开PDF
    返回: list of dict, 每个dict代表一个表格的料号信息
    格式: [{"table_type": str, "part_number": str, "material_name": str, "page_num": int}, ...]
    """
    table_infos = []

    # V5.4: 使用预提取的表格，不重新打开PDF
    if not tables:
        return table_infos

    # === V5.8.7: 表头标签关键词（用于检测和跳过表头行）===
    _header_label_keywords = {
        'supplier', 'model', 'number', 'remark', 'material', 'version',
        'description', '规格', '日期', 'date', 'rev', '料号', '物料',
        '编号', '版本', '零件', 'name', 'part', 'item', '数量', 'unit',
        '单位', '是', '否', 'rohs', 'note', '注意', 'vendor', 'tool',
        '模号', '穴数', 'cav', 'revision', 'sign', '签署', 'department',
        '部门', 'producer', '制作人', 'status', 'state', '承认'
    }

    # === V5.8.7: 非料号字段关键词（这些字段的值不应被当作料号提取）===
    _non_pn_field_labels = {
        'supplier', 'vendor', '供应商', 'tool number', '模号',
        'revision', 'rev', '版本', 'cavity', '穴数', 'date', '日期',
        'inspected', '确认者', 'comments', '签字', 'sign', 'unit', '单位',
        'supplier p/n', '供应商料号', 'supplier model', '供应商型号',
        'project name', '项目名称', 'tp fw', 'camera firmware', '摄像头'
    }

    # === V5.8.7: 非物料名称字段关键词 ===
    _non_name_field_labels = {
        'project name', '项目名称', 'tp fw', '量产版本', 'camera firmware',
        '摄像头code', 'revision', 'rev', '版本', 'vendor', '供应商',
        'tool number', '模号', 'cavity', '穴数', 'date', '日期',
        'inspected', '确认者', 'material name', '材质名称', '材质牌号',
        'material code', 'inches', '毫米', 'unit', '单位', 'comments', '签字',
        'supplier p/n', '供应商料号', 'supplier model', '供应商型号'
    }

    # === V5.8.7: 表格标题关键词（这些是整个表格的标题，不是数据）===
    _table_title_keywords = {
        'sample acknowledgement', '样品承认书', 'catalog', '目录',
        'bill of material', '物料清单', 'full size measurement', '全尺寸测量',
        'cpk report', 'cpk 报告', 'packaging method', '包装方式',
        'process flow', '制造流程图', 'qc flow', 'qc工程图',
        'rohs', 'reach', 'restricted substances', '限用物质', '调查表',
        'test report', '测试报告'
    }

    try:
        for t_dict in tables:
            table = t_dict.get("table", [])
            page_num = t_dict.get("page", 0)

            if not table or len(table) < 1:
                continue

            # 判断表格类型（基于第一个非空行的内容）— V5.4: 使用预提取表格
            table_type = "unknown"
            first_rows_text = ""
            for row in table[:3]:
                if row:
                    row_str = " ".join(str(c) if c else "" for c in row)
                    first_rows_text += row_str + " "
            _current_table_page = page_num

            first_rows_lower = first_rows_text.lower()

            if "rohs" in first_rows_lower and ("survey" in first_rows_lower or "调查表" in first_rows_text):
                table_type = "RoHS调查表"
            elif "rohs" in first_rows_lower and ("test report" in first_rows_lower or "测试报告" in first_rows_text):
                table_type = "RoHS测试报告"
            elif "reach" in first_rows_lower:
                table_type = "REACH报告"
            elif "cpk" in first_rows_lower:
                table_type = "CPK报告"
            elif "full size" in first_rows_lower or "全尺寸" in first_rows_text:
                table_type = "全尺寸测量报告"
            elif "process flow" in first_rows_lower or "制造流程图" in first_rows_text:
                table_type = "制造流程图"
            elif "sample acknowledgement" in first_rows_lower or "样品承认书" in first_rows_text:
                table_type = "样品承认书"
            elif "bill of material" in first_rows_lower or "物料清单" in first_rows_text:
                table_type = "物料清单"
            elif "packaging method" in first_rows_lower or "包装方式" in first_rows_text:
                table_type = "包装方式"

            # 从表格中提取料号和物料名称
            part_number = ""
            material_name = ""

            # === V5.8.7: 预扫描：检测哪些行是表头/标题行 ===
            header_row_indices = set()
            for ri, row in enumerate(table[:6]):
                if not row:
                    continue
                row_text_lower = " ".join(str(c).lower().strip() if c else "" for c in row)
                # 统计这一行中有多少个表头标签关键词
                label_count = sum(1 for kw in _header_label_keywords if kw in row_text_lower)
                # 如果一行有>=3个标签关键词，判定为表头行
                if label_count >= 3:
                    header_row_indices.add(ri)
                # 检查是否是表格标题行（第一行通常是表格标题）
                if ri == 0:
                    row_text_clean = " ".join(str(c).strip() if c else "" for c in row)
                    is_title_row = any(kw in row_text_clean.lower() for kw in _table_title_keywords)
                    if is_title_row:
                        header_row_indices.add(ri)
                    # 第一行如果很短且不含数字，也当作标题
                    non_empty_cells = [str(c).strip() for c in row if c and str(c).strip()]
                    if len(non_empty_cells) <= 2 and all(
                        not re.search(r'\d', cell) for cell in non_empty_cells
                    ):
                        header_row_indices.add(ri)

            # === V5.8.7: 在非表头行的数据区域搜索料号和名称 ===
            for ri, row in enumerate(table):
                # 跳过已识别的表头/标题行
                if ri in header_row_indices:
                    continue

                if not row:
                    continue

                # 构建当前行的单元格文本列表（用于跨单元格合并检测）
                cell_texts = []
                for cell in row:
                    cell_texts.append(str(cell).strip() if cell else "")

                # 尝试跨单元格合并检测料号（pdfplumber可能拆分长字符串）
                for start_ci in range(len(cell_texts)):
                    if part_number and material_name:
                        break
                    if not cell_texts[start_ci]:
                        continue

                    cell_str = cell_texts[start_ci]
                    cell_len = len(cell_str)

                    # --- 料号精确匹配 ---
                    pn_cell_patterns = [
                        r'^[A-Za-z]{1,2}\d{8,}[\w\-]*$',       # 纯料号如 K6340000520LA
                        r'^(K|M|S|NC|XC)[A-Za-z0-9_\-]{6,}$',   # 常见前缀开头的料号
                    ]
                    is_pn_cell = any(re.match(pat, cell_str) for pat in pn_cell_patterns)

                    # 如果当前单元格像部分料号（如 K6340000520），尝试与下一单元格合并
                    if not is_pn_cell and re.match(r'^(K|M)[0-9]{6,}$', cell_str):
                        # 检查下一个或几个单元格是否能拼接成完整料号
                        combined = cell_str
                        for next_ci in range(start_ci + 1, min(start_ci + 3, len(cell_texts))):
                            if cell_texts[next_ci]:
                                combined += cell_texts[next_ci]
                                if re.match(r'^[A-Za-z]{1,2}\d{8,}[\w\-]*$', combined):
                                    part_number = combined
                                    break
                            else:
                                break
                        if not part_number:
                            # 即使不能形成完整料号，部分匹配也接受
                            part_number = cell_str

                    elif is_pn_cell:
                        part_number = cell_str

                    # --- 物料名称匹配（V5.8.7严格版）---
                    if not material_name and cell_str and cell_len >= 5:
                        has_chinese = bool(re.search(r'[\u4e00-\u9fff]', cell_str))
                        has_underscore = '_' in cell_str

                        # V5.8.7: 必须满足以下条件才可能是物料名称：
                        # A) 含中文 且 (长度>=8 或 含下划线) → 如 "屏蔽盖导热硅胶 S0521_Frame..."
                        # B) 纯英文但含下划线且长度>=15 → 如 "S0521_Frame_heat_RUBBERY"
                        # C) 排除所有表格类型名、标签文字、短值
                        is_name_candidate = False
                        if has_chinese:
                            # 中文名称：至少5个汉字或混合格式
                            chinese_char_count = len(re.findall(r'[\u4e00-\u9fff]', cell_str))
                            if (chinese_char_count >= 3 or cell_len >= 8) and has_underscore:
                                is_name_candidate = True
                            elif chinese_char_count >= 5:
                                is_name_candidate = True
                        elif has_underscore and cell_len >= 15:
                            # 长下划线格式英文名
                            if re.match(r'^S\d+_[\w&]+$', cell_str, re.IGNORECASE):
                                is_name_candidate = True

                        # 排除明显不是物料名称的值
                        bad_values = {
                            'supplier', 'vendor', '小草', '/', 'v1.0', 'v1.1', 'v2.0',
                            's0521', 'k6340000520la', 'xcdp-p3135', 'xcdp-p3135',
                        }
                        cell_lower = cell_str.lower()
                        is_bad = (
                            cell_lower in bad_values or
                            re.match(r'^[A-Z]{1,2}[0-9.]+$', cell_str) or  # 纯型号如V1.0
                            any(kw in cell_lower for kw in _table_title_keywords) or
                            any(kw in cell_lower for kw in _header_label_keywords) and not has_chinese
                        )

                        # 检查同行前一个单元格是否是非名称字段的标签
                        is_non_name_field = False
                        if start_ci > 0 and cell_texts[start_ci - 1]:
                            prev_lower = cell_texts[start_ci - 1].lower().strip()
                            prev_clean = re.sub(r'[\s/（）():：]', '', prev_lower)
                            if any(nnl in prev_clean for nnl in _non_name_field_labels):
                                is_non_name_field = True

                        if is_name_candidate and not is_bad and not is_non_name_field:
                            material_name = cell_str

                if part_number and material_name:
                    break

            # 如果上面的精确模式没找到，用宽松方式搜索（但只在数据行中搜索）
            if not part_number or not material_name:
                # 只使用非表头行构建搜索文本
                data_rows_text = ""
                for ri, row in enumerate(table[:8]):
                    if ri in header_row_indices:
                        continue
                    if row:
                        row_vals = [str(c).strip() if c else "" for c in row]
                        data_rows_text += " ".join(row_vals) + " "

                # 宽松提取料号（字母+数字组合，长度>=10）
                if not part_number:
                    loose_pn = re.search(
                        r'(?:part\s*(?:number|no)|料号|零件号|零件件号|material\s*number|零件料号)[\s:：()（）\w]*[\s:：]*([A-Za-z][A-Za-z0-9_\-]{7,})',
                        data_rows_text, re.IGNORECASE
                    )
                    if loose_pn:
                        part_number = loose_pn.group(1)

                # 宽松提取物料名称（V5.8.7严格版：必须有中文>=3字 或 长下划线英文名）
                if not material_name:
                    loose_name = re.search(
                        r'(?:product\s*name|产品名称|description|零件名称|物料名称|part\s*name\s*[/／]\s*model|零件名称[/／]型号|零件名称)[\s:：()（）\w]*[\s:：]*([A-Za-z0-9_\-&+\u4e00-\u9fff]{5,})',
                        data_rows_text, re.IGNORECASE
                    )
                    if loose_name:
                        candidate = loose_name.group(1).strip()
                        # V5.8.7 严格过滤
                        has_cn = bool(re.search(r'[\u4e00-\u9fff]', candidate))
                        cn_count = len(re.findall(r'[\u4e00-\u9fff]', candidate)) if has_cn else 0
                        has_long_us = '_' in candidate and len(candidate) >= 15
                        bad_vals = {'supplier', 'vendor', 'v1.0', 'v1.1', 'v2.0', '/',
                                    '小草', 's0521', 'k6340000520la', 'xcdp-p3135'}
                        is_bad = (candidate.lower() in bad_vals or
                                 re.match(r'^[A-Z]{1,2}[0-9.]+$', candidate) or
                                 any(kw in candidate.lower() for kw in _table_title_keywords) or
                                 (not has_cn and not has_long_us))
                        if (len(candidate) >= 5 and not is_bad and
                                (cn_count >= 3 or has_long_us)):
                            material_name = candidate

            # V5.8.7: 只有提取到有效料号或有效名称时才添加结果
            # 过滤掉明显无效的结果（如表头标题、说明文字等）
            _is_valid_pn = bool(part_number and re.match(r'^[A-Za-z]{0,2}\d{6,}[\w\-]*$', part_number))
            # 名称有效性检查：必须不是表头/标题/说明文字
            _bad_name_prefixes = (
                'sample acknowledgement', '样品承认书', 'catalog', '目录',
                'bill of material', '物料清单', 'full size', '全尺寸',
                'cpk report', 'cpk 报告', 'packaging method', '包装方式',
                'product specifications', '产品规格书', 'process flow',
                'rohs', 'reach', 'restricted substances', '受控号',
                'saleable part', '可销售部件', '表单编码', 'tolerance',
                'place a layer', '箱子底', '% ', 'remark', '备注',
                '注意事项', 'material environmental', 'supplier information',
                'customer review', 'check item'
            )
            _is_valid_name = False
            if material_name:
                mn_clean = material_name.lower().strip()
                mn_first_line = material_name.split('\n')[0].strip()[:30].lower()
                _is_name_bad_prefix = any(p in mn_clean or p in mn_first_line for p in _bad_name_prefixes)
                # 有效名称：不含坏前缀 且 (含>=3个中文 或 长下划线英文名)
                cn_count = len(re.findall(r'[\u4e00-\u9fff]', material_name))
                _is_valid_name = (not _is_name_bad_prefix and
                                  (cn_count >= 3 or ('_' in material_name and len(material_name) >= 15)) and
                                  len(material_name) <= 80)

            if (_is_valid_pn or _is_valid_name):
                table_infos.append({
                    "table_type": table_type,
                    "part_number": part_number,
                    "material_name": material_name,
                    "page_num": _current_table_page,
                })

    except Exception:
        pass

    return table_infos
def check_catalog_checkboxes(pdf_path, page_analysis, tables=None):  # V5.5: 接受预提取表格
    """
    V6.2 功能1：目录/Catalog勾选状态检测
    V5.5优化: 支持从预提取的表格数据中检测勾选状态（解决表格形式目录页检测失效问题）
    判定目录中的16项是否全部已勾选
    返回: dict with checklist status
    """
    result = {
        "has_catalog_page": False,
        "catalog_page_num": 0,
        "total_items": 16,
        "checked_count": 0,
        "unchecked_count": 16,
        "checked_items": [],
        "unchecked_items": [],
        "status": "⏱ 未检测到目录",
        "issues": [],
        "details": "",
    }

    # 寻找目录页
    catalog_page = None
    for p in page_analysis:
        text = p.get("text", "")
        if ("catalog" in text.lower() or "目录" in text) and \
           ("bill of material" in text.lower() or "物料清单" in text):
            catalog_page = p
            break

    if not catalog_page:
        # 尝试用关键词找目录页
        for p in page_analysis:
            text = p.get("text", "")
            if "catalog" in text.lower() or "目录" in text:
                items_found = 0
                for kw in ["bill of material", "engineering drawing", "sample photos",
                           "full size", "cpk report", "product specification",
                           "process flow", "packaging"]:
                    if kw in text.lower():
                        items_found += 1
                if items_found >= 5:  # 至少包含5个目录项才判定为目录页
                    catalog_page = p
                    break

    if not catalog_page:
        return result

    result["has_catalog_page"] = True
    result["catalog_page_num"] = catalog_page["page_num"]
    text = catalog_page["text"]

    # ===== V5.5 新增：方法0 - 从预提取表格中检测目录勾选 =====
    _catalog_from_table = False
    if tables:
        for t_dict in tables:
            tbl = t_dict.get("table", [])
            tbl_page = t_dict.get("page", 0)
            if tbl_page != catalog_page["page_num"]:
                continue
            if not tbl or len(tbl) < 2:
                continue
            # 检查此表格是否为目录表格（含Catalog/目录标题）
            _tbl_header_text = ""
            for row in tbl[:2]:
                if row:
                    _tbl_header_text += " ".join(str(c) if c else "" for c in row) + " "
            if "catalog" in _tbl_header_text.lower() or "目录" in _tbl_header_text:
                # 这是目录表格！逐行分析勾选状态
                _checked_indicators = ['☑', '☒', '✓', '✔', '✅', '[x]', '[X]', '(x)', '(X)', '√', 'checked', 'yes']
                _unchecked_indicators = ['☐', '□', '✗', '✖', '❌', '[ ]', '( )', '○', 'unchecked', 'no']
                _t_checked = 0
                _t_unchecked = 0
                _t_items = []
                for row in tbl[1:]:  # 跳过表头行
                    if not row:
                        continue
                    row_text = " ".join(str(c) if c else "" for c in row)
                    if not row_text.strip():
                        continue
                    # 判断此行的勾选状态
                    _has_chk = any(ind in row_text for ind in _checked_indicators)
                    _has_unchk = any(ind in row_text for ind in _unchecked_indicators)
                    item_label = row_text[:60].strip()
                    if _has_chk:
                        _t_checked += 1
                        _t_items.append({"item": item_label, "status": "✅ 已勾选"})
                        result["checked_items"].append(item_label)
                    else:
                        _t_unchecked += 1
                        _t_items.append({"item": item_label, "status": "❌ 未勾选"})
                        result["unchecked_items"].append(item_label)
                if _t_checked + _t_unchecked >= 3:  # 至少检测到3个目录项才采用表格结果
                    result["checked_count"] = _t_checked
                    result["unchecked_count"] = _t_unchecked
                    result["total_items"] = _t_checked + _t_unchecked
                    _catalog_from_table = True
                    
                    # V5.7增强：检查是否存在任何checkbox标记字符
                    # 如果完全没有任何checkbox标记（☑/☐/✓/□等），说明这是普通列表式目录而非勾选式
                    _has_any_checkbox = False
                    for row in tbl[1:]:
                        if not row:
                            continue
                        row_text = " ".join(str(c) if c else "" for c in row)
                        _chk_chars = ['☑', '☐', '✓', '✔', '✅', '❏', '√', '[x]', '[X]', '[ ]', '(x)', '( )']
                        if any(ch in row_text for ch in _chk_chars):
                            _has_any_checkbox = True
                            break
                    
                    # 直接跳转到结果判定
                    if _has_any_checkbox:
                        # 有checkbox标记：按实际勾选状态判断
                        if _t_unchecked == 0 and _t_checked > 0:
                            result["status"] = "✅ 全部已勾选"
                            result["details"] = f"目录共{_t_checked}项，全部已勾选（表格检测）"
                        elif _t_unchecked == result["total_items"]:
                            result["status"] = "❌ NG - 目录全部未勾选"
                            result["details"] = f"目录共{result['total_items']}项，**全部未勾选**（第{result['catalog_page_num']}页）"
                            result["issues"].append(f"目录页（第{result['catalog_page_num']}页）：{result['total_items']}项全部未勾选")
                        elif _t_unchecked > 0:
                            result["status"] = f"⚠️ 部分未勾选（{_t_unchecked}/{result['total_items']}）"
                            result["details"] = f"目录共{result['total_items']}项，已勾选{_t_checked}项，未勾选{_t_unchecked}项"
                            result["issues"].append(f"目录页有{_t_unchecked}项未勾选")
                    else:
                        # V5.7：无checkbox标记 → 判定为目录已提供（列表式目录）
                        result["status"] = "✅ 目录已提供"
                        result["details"] = f"目录共{result['total_items']}项（列表式目录，无勾选标记框）"
                        result["checked_count"] = result["total_items"]  # 视为全部有效
                        result["unchecked_count"] = 0
                    return result

    # 方法1：通过文本特征判断勾选状态
    # PDF中的checkbox可能以 ☑(checked) / ☐(unchecked) / ✓ / □ / [x] / [ ] 等形式出现
    # 也可能是 Unicode checkbox 字符

    # 统计目录项数量（按序号1-16）
    catalog_items = []
    item_pattern = r'(?:^|\n)\s*(\d{1,2})\s+(.{10,60})'

    # 更健壮的方法：查找所有目录项文本
    expected_items = [
        "Bill of material", "Engineering drawings", "Sample photos",
        "Full size measurement report", "Cpk Report", "Product specifications",
        "Process Flow Chart", "Packaging method", "QC Flow Chart",
        "Electrical performance test report", "Reliability test report",
        "Material certificate", "RoHS 2.0 restricted substances composition questionnaire",
        "RoHS 2.0 Test Report", "REACH", "REACH Test Report"
    ]
    expected_cn_items = ["物料清单", "工程图纸", "样品照片", "全尺寸测量报告",
                         "Cpk报告", "产品规格书", "制造流程图", "包装方式",
                         "QC工程图", "电气性能测试报告", "可靠性测试报告",
                         "材质证明", "RoHS", "REACH"]

    checked_indicators = ['☑', '☒', '✓', '✔', '✅', '[x]', '[X]', '(x)', '(X)', '√']
    unchecked_indicators = ['☐', '□', '✗', '✖', '❌', '[ ]', '( )', '○']

    # 分析每一行的勾选状态
    lines = text.split('\n')
    checked = 0
    unchecked = 0
    item_details = []

    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue

        # 检查是否为目录项行（含序号+项目名称）
        is_catalog_item = False
        item_name = ""

        for item in expected_items + expected_cn_items:
            if item.lower() in line_stripped.lower():
                is_catalog_item = True
                item_name = item
                break

        # 也匹配数字序号格式
        if not is_catalog_item:
            num_match = re.match(r'^(\d{1,2})\s+', line_stripped)
            if num_match:
                is_catalog_item = True
                item_name = line_stripped[:50]

        if is_catalog_item:
            # 判断此行的勾选状态
            has_checked = any(ind in line for ind in checked_indicators)
            has_unchecked = any(ind in line for ind in unchecked_indicators)

            # 特殊处理：如果同时没有明确的checked/unchecked标记，
            # 尝试从行的末尾或特定位置判断
            if not has_checked and not has_unchecked:
                # PDF中未勾选的checkbox常显示为 ☐ 或空方框
                # 已勾选的可能有 ✓ 或其他标记
                if '☐' in line or '□' in line:
                    has_unchecked = True
                elif '☑' in line:
                    has_checked = True
                else:
                    # 无法确定，默认为未勾选（保守策略）
                    has_unchecked = True

            if has_checked:
                checked += 1
                item_details.append({"item": item_name, "status": "✅ 已勾选"})
                result["checked_items"].append(item_name)
            else:
                unchecked += 1
                item_details.append({"item": item_name, "status": "❌ 未勾选"})
                result["unchecked_items"].append(item_name)

    result["checked_count"] = checked
    result["unchecked_count"] = unchecked
    result["total_items"] = checked + unchecked if (checked + unchecked) > 0 else 16

    # 判定结果
    if result["total_items"] == 0:
        result["status"] = "⏱ 未检测到目录项"
        result["details"] = "未能解析目录中的勾选框，建议人工确认"
    elif unchecked == 0 and checked > 0:
        result["status"] = "✅ 全部已勾选"
        result["details"] = f"目录共{checked}项，全部已勾选"
    elif unchecked == result["total_items"]:
        # 全部未勾选 → NG
        result["status"] = "❌ NG - 目录全部未勾选"
        result["details"] = f"目录共{result['total_items']}项，**全部未勾选**（第{result['catalog_page_num']}页）"
        result["issues"].append(f"目录页（第{result['catalog_page_num']}页）：{result['total_items']}项全部未勾选")
    elif unchecked > 0:
        # 部分未勾选 → 警告
        result["status"] = f"⚠️ 部分未勾选（{unchecked}/{result['total_items']}）"
        result["details"] = f"目录共{result['total_items']}项，已勾选{checked}项，未勾选{unchecked}项"
        result["issues"].append(f"目录页有{unchecked}项未勾选：{'、'.join(result['unchecked_items'][:5])}...")

    return result


def check_part_number_consistency(page_analysis, pdf_path, tables=None):  # V5.4: 接受预提取tables
    """
    V6.2 功能2：料号&物料名称跨表一致性检查
    提取封面和各表头的料号/物料名称，比对一致性
    V5.4优化: 接受预提取的tables参数，避免重新打开PDF
    返回: dict with consistency results
    """
    result = {
        "cover_info": {},
        "table_infos": [],
        "consistency_checks": [],
        "overall_status": "✅ 一致",
        "issues": [],
    }

    # Step 1: 提取封面的料号和物料名称
    cover = extract_cover_info(page_analysis, pdf_path, tables=tables)
    result["cover_info"] = cover

    if not cover["part_number"] and not cover["material_name"]:
        result["overall_status"] = "⏱ 无法提取封面信息"
        result["issues"].append("无法从封面/样品承认书页提取料号或物料名称")
        return result

    # Step 2: 提取各表头的料号和物料名称
    table_infos = extract_table_headers_part_info(page_analysis, tables=tables)  # V5.4: 使用预提取表格
    result["table_infos"] = table_infos

    # V5.8.5/V5.8.7：过滤掉外部检测报告、调查表和非关键页面，只对比文档内部核心表格
    # 外部报告/调查表的料号格式可能与内部不同（如Supplier P/N），不具备可比性
    _external_report_types = {
        "RoHS测试报告", "RoHS调查表", "REACH报告", "CPK报告",
        "SGS报告", "华测报告", "材质证明",
        # V5.8.7: 以下页面类型不参与料号一致性对比
        "包装方式", "样品照片", "unknown",  # unknown类型通常是非标准格式页面
    }
    _internal_table_infos = [
        ti for ti in table_infos if ti.get("table_type", "unknown") not in _external_report_types
    ]
    # 同时记录被排除的外部报告（用于信息展示）
    _excluded_count = len(table_infos) - len(_internal_table_infos)

    if not _internal_table_infos:
        result["overall_status"] = "⚠️ 未检测到其他表格的表头信息"
        result["issues"].append("未在各报告表头中找到料号/物料名称信息（已排除外部检测报告）")
        return result

    # Step 3: 逐一比对（仅对比内部表格）
    ref_pn = cover["part_number"]
    ref_name = cover["material_name"]

    for ti in _internal_table_infos:
        table_type = ti["table_type"]
        table_pn = ti["part_number"]
        table_name = ti["material_name"]
        page_num = ti["page_num"]

        check_result = {
            "table_type": table_type,
            "page_num": page_num,
            "pn_match": None,
            "name_match": None,
            "pn_detail": "",
            "name_detail": "",
        }

        # 料号比对
        if ref_pn and table_pn:
            # 标准化比较（忽略大小写和空格）
            pn_normalized_ref = ref_pn.upper().replace(" ", "").replace("-", "")
            pn_normalized_table = table_pn.upper().replace(" ", "").replace("-", "")
            if pn_normalized_ref == pn_normalized_table:
                check_result["pn_match"] = "✅ 一致"
                check_result["pn_detail"] = f"{table_pn} == {ref_pn}"
            else:
                check_result["pn_match"] = "❌ 不一致"
                check_result["pn_detail"] = f"表头:{table_pn} ≠ 封面:{ref_pn}"
                result["issues"].append(f"[{table_type}(第{page_num}页)] 料号不一致：表头'{table_pn}' ≠ 封面'{ref_pn}'")
        elif ref_pn and not table_pn:
            check_result["pn_match"] = "⏱ 表头无料号"
            check_result["pn_detail"] = f"封面料号:{ref_pn}，表头未找到料号"
        else:
            check_result["pn_match"] = "⏱ 无参考"
            check_result["pn_detail"] = ""

        # 物料名称比对
        if ref_name and table_name:
            # V5.8.7: 标准化比较（忽略大小写、空格、换行、全角半角差异）
            import re as _re
            name_normalized_ref = _re.sub(r'[\s\n\r\t－_]', '', ref_name.upper())
            name_normalized_table = _re.sub(r'[\s\n\r\t－_]', '', table_name.upper())

            # 允许微小差异（如全尺寸报告多了"线"字）
            if name_normalized_ref == name_normalized_table:
                check_result["name_match"] = "✅ 一致"
                check_result["name_detail"] = f"{table_name} == {ref_name}"
            elif name_normalized_ref in name_normalized_table or name_normalized_table in name_normalized_ref:
                check_result["name_match"] = "⚠️ 基本一致（微小差异）"
                diff = ""
                if len(name_normalized_ref) != len(name_normalized_table):
                    diff = f"(长度差{len(name_normalized_ref)} vs {len(name_normalized_table)})"
                check_result["name_detail"] = f"表头:'{table_name}' ≈ 封面:'{ref_name}' {diff}"
                result["issues"].append(f"[{table_type}(第{page_num}页)] 物料名称存在微小差异：表头'{table_name}' ≠ 封面'{ref_name}'")
            else:
                check_result["name_match"] = "❌ 不一致"
                check_result["name_detail"] = f"表头:{table_name} ≠ 封面:{ref_name}"
                result["issues"].append(f"[{table_type}(第{page_num}页)] 物料名称不一致：表头'{table_name}' ≠ 封面'{ref_name}'")
        elif ref_name and not table_name:
            check_result["name_match"] = "⏱ 表头无物料名称"
            check_result["name_detail"] = f"封面物料名称:{ref_name}，表头未找到"
        else:
            check_result["name_match"] = "⏱ 无参考"
            check_result["name_detail"] = ""

        result["consistency_checks"].append(check_result)

    # Step 4: V5.9.1 新增 —— 文件名料号 vs 文档内容料号一致性检查
    # 用户场景：文件命名为 CRS_K5350000042LA_...pdf，但文档内料号为 K5311000042LA，
    # 这类"文不对题"的命名错误需在料号一致性检查中暴露出来
    _fname = os.path.basename(pdf_path)
    _fn_pn_match = re.search(r'(K\d{6,}[A-Za-z]*)', _fname, re.IGNORECASE)
    _fn_pn = None
    if _fn_pn_match:
        _fn_pn = _fn_pn_match.group(1).upper().replace(" ", "").replace("-", "")
    result["filename_part_number"] = _fn_pn  # 供UI/Excel展示

    if _fn_pn:
        # 4.1 与封面料号比对
        if ref_pn:
            _ref_pn_norm = ref_pn.upper().replace(" ", "").replace("-", "")
            if _fn_pn != _ref_pn_norm:
                result["issues"].append(
                    f"[文件名 vs 封面] 料号不一致：文件名'{_fn_pn}' ≠ 封面料号'{ref_pn}'"
                )
                result["consistency_checks"].append({
                    "table_type": "文件名 vs 封面",
                    "page_num": 0,
                    "pn_match": "❌ 不一致",
                    "name_match": None,
                    "pn_detail": f"文件名:{_fn_pn} ≠ 封面:{ref_pn}",
                    "name_detail": "",
                })

        # 4.2 与内部各表格料号比对（封面提取失败时可在此发现差异）
        for ti in _internal_table_infos:
            _tp = ti.get("part_number")
            if _tp:
                _tp_norm = _tp.upper().replace(" ", "").replace("-", "")
                if _fn_pn != _tp_norm:
                    result["issues"].append(
                        f"[文件名 vs {ti.get('table_type', '未知')}(第{ti.get('page_num', 0)}页)] "
                        f"料号不一致：文件名'{_fn_pn}' ≠ 表头'{_tp}'"
                    )
                    result["consistency_checks"].append({
                        "table_type": f"文件名 vs {ti.get('table_type', '未知')}",
                        "page_num": ti.get("page_num", 0),
                        "pn_match": "❌ 不一致",
                        "name_match": None,
                        "pn_detail": f"文件名:{_fn_pn} ≠ 表头:{_tp}",
                        "name_detail": "",
                    })

    # V5.8.3 修复：判定整体状态
    # 料号一致性检查的核心是**料号(Part Number)**的一致性，物料名称仅作参考
    # 只有料号不一致才判定为❌不合格；物料名称不一致仅记录为⚠️警告
    has_pn_mismatch = any("❌" in str(c.get("pn_match", "")) for c in result["consistency_checks"])
    has_pn_minor = any("⚠️" in str(c.get("pn_match", "")) for c in result["consistency_checks"])
    has_name_mismatch = any("❌" in str(c.get("name_match", "")) for c in result["consistency_checks"])
    has_name_minor = any("⚠️" in str(c.get("name_match", "")) for c in result["consistency_checks"])

    # 将物料名称的❌降级为⚠️警告（物料名称在不同表格中可能有合理差异）
    if has_name_mismatch:
        for issue in list(result["issues"]):
            if "物料名称" in issue or "产品名称" in issue:
                result["issues"].remove(issue)
                result["issues"].append(issue.replace("❌", "⚠️").replace("不一致", "存在差异(参考)"))

    if has_pn_mismatch:
        result["overall_status"] = "❌ 存在不一致"
    elif has_pn_minor or has_name_mismatch:
        result["overall_status"] = "⚠️ 基本一致（有微小差异）"
    else:
        result["overall_status"] = "✅ 全部一致"

    return result


def generate_final_verdict_v62(material_type, all_results, standards):
    """
    V6.2 第六类：生成最终检验结论与处理建议（含目录勾选+料号一致性）
    """
    issues = []
    completeness = all_results.get("completeness", {})
    rohs = all_results.get("rohs", {})
    cpk = all_results.get("cpk", {})
    dimension = all_results.get("dimension", {})
    validity = all_results.get("validity", {})

    issues.extend(completeness.get("issues", []))
    issues.extend(rohs.get("issues", []))
    issues.extend(cpk.get("issues", []))
    # V5.9: 尺寸对应性作为参考项，不加入强制issues列表（需人工核对）
    # issues.extend(dimension.get("issues", []))
    issues.extend(validity.get("issues", []))

    # V6.2: 新增目录勾选和料号一致性问题
    catalog_check = all_results.get("catalog_check", {})
    part_consistency = all_results.get("part_consistency", {})

    if catalog_check.get("issues"):
        issues.extend(catalog_check["issues"])

    if part_consistency.get("issues"):
        issues.extend(part_consistency["issues"])

    # V5.9.0: 螺丝类物料图纸特殊要求问题
    screw_check = all_results.get("screw_check", {})
    if screw_check.get("is_screw") and screw_check.get("issues"):
        issues.extend(screw_check["issues"])

    # V5.9.2: 封面供应商信息完整性问题
    supplier_check = all_results.get("supplier_check", {})
    if supplier_check.get("issues"):
        issues.extend(supplier_check["issues"])

    total_fail = completeness.get("fail_count", 0)
    critical_fail = (
        completeness["status"] == "❌ 不合格"
        or rohs["overall_status"] == "❌ 不合格"
        or cpk["overall_status"] == "❌ 不合格"
        # V5.9: 尺寸对应性作为参考项，不影响最终合格/不合格判定
        # or dimension["overall_status"] == "❌ 不合格"
        or validity["overall_status"] == "❌ 不合格"
        or catalog_check.get("status", "").startswith("❌")
        or part_consistency.get("overall_status", "").startswith("❌")
        or (screw_check.get("is_screw") and screw_check.get("overall_status", "").startswith("❌"))  # V5.9.0
        or supplier_check.get("overall_status", "").startswith("❌")  # V5.9.2
    )

    if critical_fail or len(issues) > 3:
        verdict = "❌ 不合格，退回重报"
        suggestion = "退回供应商，要求按XC-R-0802-DQM-002正式样品承认书模板重新提交完整的样品承认书\n\n**必须补充/修正以下问题：**\n"
        for item in completeness.get("items", []):
            if "❌" in item["结果"]:
                suggestion += f"- ❌ **{item['项目']}**\n"

        # V6.2: 目录勾选NG
        if catalog_check.get("status", "").startswith("❌"):
            suggestion += f"- ❌ **目录全部未勾选**（第{catalog_check['catalog_page_num']}页）：请在样品承认书目录中勾选所有已提供的文件项\n"
        elif catalog_check.get("unchecked_count", 0) > 0 and catalog_check.get("unchecked_count") < catalog_check.get("total_items", 16):
            suggestion += f"- ⚠️ **目录部分未勾选**（{catalog_check['unchecked_count']}/{catalog_check['total_items']}项未勾选）\n"

        for issue in issues:
            suggestion += f"- ⚠️ {issue}\n"
    elif len(issues) > 0:
        verdict = "⚠️ 基本合格，需补充材料"
        suggestion = "该封样报告基本符合要求，但存在以下问题需补充或修正：\n\n"
        for i, issue in enumerate(issues, 1):
            suggestion += f"{i}. {issue}\n"
    else:
        verdict = "✅ 合格，建议通过"
        suggestion = "该封样报告符合XC-R-0802-DQM-002模板要求，建议通过审核。"

    return {
        "verdict": verdict,
        "issues": issues,
        "issue_count": len(issues),
        "suggestion": suggestion,
        "material_type": material_type,
    }


# 保持向后兼容
def generate_final_verdict(material_type, all_results, standards):
    """向后兼容的包装"""
    return generate_final_verdict_v62(material_type, all_results, standards)


# ============================================================
# 主审核流程（V4.0 完整8步）
# ============================================================

def _build_detail_str(sub_items, extra_values):
    """
    V5.8.3: 从sub_items和额外值构建详情字符串（用于Excel汇总表）
    将检查项的子结果合并为可读的摘要字符串
    """
    parts = []
    if sub_items:
        for k, v in list(sub_items.items())[:5]:  # 最多取前5个子项
            if isinstance(v, str) and v.strip():
                # 截断过长的值
                short_v = v[:50] + "..." if len(v) > 50 else v
                parts.append(f"{short_v}")
    if extra_values:
        parts.append(f"值:{extra_values}")
    return " | ".join(parts) if parts else ""


def run_full_inspection(file_path, file_name, standards):
    """执行完整的V4.0 6大类审核流程（含PDF逐页分析）"""
    check_date = datetime.now()
    version = standards.get("version", "未知")

    # 提取PDF并逐页分析（V4.0 核心）—— V5.2 优化：一次打开完成所有解析
    page_analysis, all_text, tables = analyze_pdf_page_by_page(file_path)

    # 第零步：物料类型判定 —— V5.5: 提取前3页文本作为内容提示辅助识别
    _mat_hint = ""
    for p in page_analysis[:min(3, len(page_analysis))]:
        _mat_hint += p.get("text", "") + " "
    mat_type, mat_type_cn, mat_kw, needs_elec = determine_material_type(
        file_name, standards, pdf_content_hint=_mat_hint.strip()
    )

    # 第一步：文件类型检查（V4.0新增）
    file_type, file_type_note = check_file_type(file_path, all_text)

    # 第二步~第五步：各类检验（基于逐页分析结果）
    completeness = inspect_file_completeness_v4(page_analysis, mat_type, standards)
    rohs = inspect_rohs_compliance(page_analysis, standards, check_date, tables=tables)
    cpk = inspect_cpk_compliance(page_analysis, standards, all_text, tables=tables)
    dimension = inspect_dimension_correspondence(page_analysis, standards)
    validity = inspect_report_validity(page_analysis, standards, check_date)

    # V6.2 新增：目录勾选状态检测 + 料号跨表一致性检查
    catalog_check = check_catalog_checkboxes(file_path, page_analysis, tables=tables)  # V5.5: 传入预提取表格
    part_consistency = check_part_number_consistency(page_analysis, file_path, tables=tables)  # V5.4: 传入预提取表格

    # V5.9.0 新增：螺丝/紧固件类物料工程图纸特殊要求检查
    _cover_name = part_consistency.get("cover_info", {}).get("material_name", "")
    screw_check = check_screw_drawing_requirements(page_analysis, tables=tables, material_name=_cover_name)

    # V5.9.2 新增：封面供应商信息完整性检查
    supplier_check = check_supplier_info_completeness(page_analysis, tables=tables)

    # V5.7 新增：根据料号识别物料类型（基于物料编码规则）
    _coding_rules = load_material_coding_rules()
    _mat_type_detail = mat_type_cn  # 默认使用原来的识别结果
    if _coding_rules:
        _pn = part_consistency.get("cover_info", {}).get("part_number", "")
        if not _pn:
            # 如果封面没提取到，尝试从文件名提取
            _pn_match = re.search(r"([A-Za-z]\d{4}\d+[A-Za-z]*)", file_name)
            if _pn_match:
                _pn = _pn_match.group(1)
        if _pn:
            _mat_info = identify_material_type(_pn, _coding_rules)
            if _mat_info["成功"]:
                _mat_type_detail = _mat_info["详情"]
            else:
                _mat_type_detail = f"{mat_type_cn}（编码规则：{_mat_info['详情']}）"


    # 第六步：生成最终结论
    all_results = {
        "completeness": completeness,
        "rohs": rohs,
        "cpk": cpk,
        "dimension": dimension,
        "validity": validity,
        "catalog_check": catalog_check,
        "part_consistency": part_consistency,
        "screw_check": screw_check,  # V5.9.0
        "supplier_check": supplier_check,  # V5.9.2
    }
    final = generate_final_verdict_v62(mat_type, all_results, standards)

    # 构建返回结果
    result = {
        "文件名": file_name,
        "文件类型": file_type_note,
        "物料类型": _mat_type_detail if "_mat_type_detail" in dir() else (f"{mat_type_cn}" if mat_type != "unknown" else "未知（需人工确认）"),
        "需要电气性能测试": "是" if needs_elec else "否",
        "文件完整性": completeness["status"],
        "RoHS合规性": rohs["overall_status"],
        "CPK合规性": cpk["overall_status"],
        # V5.8.3: 增加CPK详情
        "CPK详情": _build_detail_str(cpk.get("sub_items", {}), cpk.get("cpk_values", [])),
        "尺寸对应性": dimension["overall_status"],
        # V5.8.3: 增加尺寸详情
        "尺寸详情": _build_detail_str(dimension.get("sub_items", {}), []),
        "报告时效性": validity["overall_status"],
        # V5.8.3: 增加报告时效详情
        "时效详情": _build_detail_str(validity.get("sub_items", {}), []),
        # V6.2 新增
        "目录勾选状态": catalog_check.get("status", "⏱ 未检测"),
        "料号一致性": part_consistency.get("overall_status", "⏱ 未检测"),
        # V5.9.0 新增：螺丝类物料图纸特殊要求
        "螺丝图纸要求": screw_check.get("overall_status", "⏭️ 不适用") if screw_check.get("is_screw") else "⏭️ 不适用",
        # V5.9.2 新增：封面供应商信息完整性
        "供应商信息": supplier_check.get("overall_status", "⏱ 未检测"),
        # 原有
        "总体结论": final["verdict"],
        "问题数量": final["issue_count"],
        "审核时间": check_date.strftime("%Y-%m-%d %H:%M:%S"),
        "标准版本": f"V{version}",
        "_detail": {
            "mat_type": mat_type,
            "mat_type_cn": _mat_type_detail if "_mat_type_detail" in dir() else mat_type_cn,
            "needs_elec": needs_elec,
            "file_type": file_type,
            "file_type_note": file_type_note,
            "page_analysis": [
                {
                    "page": p["page_num"],
                    "is_drawing": p.get("is_engineering_drawing"),
                    "drawing_number": p.get("drawing_number"),
                    "is_bom": p.get("is_bom"),
                    "is_cpk": p.get("is_cpk"),
                    "is_rohs": p.get("is_rohs_survey"),
                    "is_reach": p.get("is_reach_survey"),
                }
                for p in page_analysis if "error" not in p
            ],
            "completeness": completeness,
            "rohs": rohs,
            "cpk": cpk,
            "dimension": dimension,
            "validity": validity,
            # V6.2 新增
            "catalog_check": catalog_check,
            "part_consistency": part_consistency,
            "screw_check": screw_check,  # V5.9.0
            "supplier_check": supplier_check,  # V5.9.2
            #
            "final": final,
        },
    }
    return result


# ============================================================
# Streamlit UI
# ============================================================

# 设置页面配置
st.set_page_config(
    page_title="封样检验应用",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)



# ============================================================
# V5.8 出错复位机制 - Session State 初始化
# ============================================================
if "app_error" not in st.session_state:
    st.session_state["app_error"] = None
if "inspection_running" not in st.session_state:
    st.session_state["inspection_running"] = False

# 加载标准
standards = load_standards()
if standards is None:
    st.error("❌ 无法加载审核标准文件（inspection_standards.json），请确认文件存在")
    st.stop()

version = standards.get("version", "未知")
last_updated = standards.get("last_updated", "未知")

# 标题区
st.title("📋 封样检验应用")
st.caption(f"基于 XC-R-0802-DQM-002 物料正式样品承认书模板 | 标准版本 V{version} | 更新于 {last_updated}")
st.markdown("---")

# 侧边栏设置
st.sidebar.header("⚙️ 审核标准设置")

# 从标准文件动态生成选项
electronic_items = standards.get("file_completeness", {}).get("electronic", {}).get("items", [])
all_item_names = [item["name"] for item in electronic_items]
default_names = ["物料清单", "工程图纸", "样品照片", "全尺寸测量报告", "Cpk报告",
                 "产品规格书", "制造流程图", "包装方式", "QC工程图", "可靠性测试报告",
                 "材质证明", "RoHS 2.0限用物质成分调查表", "RoHS 2.0测试报告"]

st.sidebar.subheader("文件完整性检查")
check_list = st.sidebar.multiselect(
    "选择要检查的项目",
    all_item_names,
    default=[n for n in default_names if n in all_item_names],
)

st.sidebar.subheader("专项检查")
rohs_check = st.sidebar.checkbox("✅ RoHS合规性检验（含日期+红框字段）", value=True)
cpk_check = st.sidebar.checkbox("✅ CPK合规性检验（≥1.33）", value=True)
dim_check = st.sidebar.checkbox("✅ 尺寸公差对应性检验", value=True)
validity_check = st.sidebar.checkbox("✅ 报告时效性检验（≤1年）", value=True)
# V6.2 新增
catalog_check = st.sidebar.checkbox("✅ 目录勾选状态检测（V6.2）", value=True)
part_check = st.sidebar.checkbox("✅ 料号&物料名称跨表一致性（V6.2）", value=True)

# ============================================================
# V5.8 复位按钮
# ============================================================
st.sidebar.markdown("---")
if st.sidebar.button("🔄 重置应用", type="secondary", use_container_width=True, help="清除所有会话状态，重新开始"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

if st.session_state.get("app_error"):
    st.sidebar.error("⚠️ 上次运行出错")
    if st.sidebar.button("🚨 清除错误状态", type="secondary"):
        st.session_state["app_error"] = None
        st.rerun()


min_cpk_display = 1.33
st.sidebar.info(f"当前CPK合格标准：≥{min_cpk_display}")

# 主界面两栏布局
col1, col2 = st.columns([1, 1])

with col1:
    st.header("📤 上传PDF文件")

    uploaded_files = st.file_uploader(
        "选择PDF文件（支持批量上传 / 拖拽上传）",
        type=["pdf"],
        accept_multiple_files=True,
        help="可一次选多个PDF文件，或直接拖入此区域"
    )

    st.subheader("或者使用文件夹路径扫描")
    folder_path = st.text_input(
        "输入包含PDF文件的文件夹路径",
        placeholder="例如：D:\\封样\\S1651\\楚鑫"
    )

    if st.button("📁 扫描文件夹", type="secondary"):
        if folder_path and os.path.exists(folder_path):
            pdf_list = []
            for root, dirs, files in os.walk(folder_path):
                for fn in files:
                    if fn.lower().endswith(".pdf"):
                        pdf_list.append(os.path.join(root, fn))
            if pdf_list:
                st.success(f"✅ 找到 **{len(pdf_list)}** 个PDF文件")
                st.session_state['folder_files'] = pdf_list
            else:
                st.warning("⚠️ 该文件夹中没有PDF文件")
        else:
            st.error("❌ 路径不存在")

    # 显示已选文件列表
    if uploaded_files or 'folder_files' in st.session_state:
        st.subheader("📄 已选择的文件")
        count = 0
        total_size_mb = 0

        if uploaded_files:
            for f in uploaded_files:
                size_mb = len(f.getvalue()) / (1024 * 1024)
                total_size_mb += size_mb
                size_warning = " ⚠️ 大文件" if size_mb > 20 else ""
                st.text(f"✅ {f.name} ({size_mb:.1f}MB){size_warning}")
                count += 1
        if 'folder_files' in st.session_state:
            for fp in st.session_state['folder_files']:
                try:
                    total_size_mb += os.path.getsize(fp) / (1024 * 1024)
                except OSError:
                    pass
                st.text(f"✅ {os.path.basename(fp)}")
                count += 1

        st.info(f"共 **{count}** 个文件待审核 | 总大小 **{total_size_mb:.1f}MB**")

        # V5.1: 大文件/多文件警告
        if total_size_mb > 50 or count > 5:
            st.warning("⚠️ 文件较大或数量较多，建议分批处理以避免超时。单个文件超过20MB可能需要较长处理时间。")

        if st.button("🗑️ 清空文件列表", type="secondary"):
            if 'folder_files' in st.session_state:
                del st.session_state['folder_files']
            st.rerun()

# ============================================================
# V5.8 错误状态检查 - 如果之前出错，显示友好错误页
# ============================================================
_app_error = st.session_state.get("app_error")
if _app_error:
    st.error("❌ 应用运行出错")
    with st.expander("查看错误详情", expanded=False):
        st.code(_app_error[:3000], language="text")
    col_reset1, col_reset2 = st.columns(2)
    with col_reset1:
        if st.button("🔄 重置应用并重新开始", type="primary", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    with col_reset2:
        if st.button("📋 复制错误信息", type="secondary", use_container_width=True):
            st.code(_app_error[:500], language="text")
            st.info("请复制上方错误信息并联系开发者")
    st.stop()

with col2:
    st.header("📊 审核结果")

    run_btn = st.button("🚀 开始审核", type="primary", use_container_width=True)

    if run_btn:


        all_paths = []
        if uploaded_files:
            seen_names = set()  # V5.3: 去重用
            for uf in uploaded_files:
                fname = uf.name
                # V5.3: 按文件名去重（同名文件只保留最后一个）
                if fname in seen_names:
                    st.warning(f"⚠️ 跳过重复文件: {fname}")
                    continue
                seen_names.add(fname)
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
                tmp.write(uf.getvalue())
                all_paths.append((tmp.name, fname))
        if 'folder_files' in st.session_state:
            seen_folder = set()  # V5.3: 文件夹扫描也去重
            for fp in st.session_state['folder_files']:
                bname = os.path.basename(fp)
                if bname in seen_folder:
                    continue
                seen_folder.add(bname)
                all_paths.append((fp, bname))

        if not all_paths:
            st.warning("⚠️ 请先上传或选择PDF文件")
        else:
            st.info(f"开始审核 **{len(all_paths)}** 个文件... (V5.9.2: 新增封面供应商信息完整性检查)")

            progress = st.progress(0)
            detail_results = []
            status_text = st.empty()  # V5.1: 实时状态更新容器

            for i, (fp, fname) in enumerate(all_paths):
                # V5.1: 实时更新当前处理状态
                status_text.info(f"📄 正在审核 [{i+1}/{len(all_paths)}]: **{fname}**")

                try:
                    res = run_full_inspection(fp, fname, standards)
                    detail_results.append(res)

                    # V5.1: 处理成功后立即释放该文件相关内存
                    if '_detail' in res:
                        res['_detail'].pop('page_analysis', None)  # 释放页面分析数据
                        for key in ['completeness', 'rohs', 'cpk', 'dimension', 'validity']:
                            if key in res['_detail']:
                                sub = res['_detail'][key]
                                if isinstance(sub, dict):
                                    sub.pop('items', None)  # 释放详细项目列表

                except Exception as e:
                    import traceback
                    error_msg = str(e)
                    st.error(f"❌ 审核文件 {fname} 时出错: {error_msg}")
                    # V5.8: 记录错误状态，供复位机制使用
                    if not st.session_state.get("app_error"):
                        st.session_state["app_error"] = f"文件审核异常({type(e).__name__}): {error_msg}\n{fname}"
                    # V5.3: 补全_detail所有字段，避免后续展示时KeyError
                    detail_results.append({
                        "文件名": fname,
                        "总体结论": f"❌ 审核异常: {error_msg[:80]}",
                        "问题数量": 1,
                        "_detail": {
                            "final": {
                                "verdict": f"❌ 异常",
                                "suggestion": f"审核出错({type(e).__name__}): {str(e)[:200]}",
                                "issue_count": 1,
                                "error_traceback": traceback.format_exc()[-500:]
                            },
                            # V5.3: 补全所有可能被展示访问的键
                            "mat_type": "unknown",
                            "mat_type_cn": "未知(异常)",
                            "needs_elec": False,
                            "file_type": "unknown",
                            "file_type_note": "⏱ 异常中断",
                            "page_analysis": [],
                            "completeness": {"items": [], "status": "⏱ 异常", "pass_count": 0, "fail_count": 0},
                            "rohs": {"sub_items": {}, "overall_status": "⏱ 异常", "issues": []},
                            "cpk": {"sub_items": {}, "overall_status": "⏱ 异常", "issues": [], "cpk_values": []},
                            "dimension": {"sub_items": {}, "overall_status": "⏱ 异常", "issues": []},
                            "validity": {"sub_items": {}, "overall_status": "⏱ 异常", "issues": []},
                            "catalog_check": None,
                            "part_consistency": None,
                            "screw_check": None,  # V5.9.0
                            "supplier_check": None,  # V5.9.2
                        },
                        **{k: "⏱ 异常" for k in [
                            "文件类型","物料类型","需要电气性能测试",
                            "文件完整性","RoHS合规性","CPK合规性",
                            "尺寸对应性","报告时效性","目录勾选状态","料号一致性",
                            "螺丝图纸要求",  # V5.9.0
                            "供应商信息",  # V5.9.2
                            "审核时间","标准版本"
                        ]}
                    })

                progress.progress((i + 1) / len(all_paths))

                # V5.1: 每处理完一个文件，显式回收内存（关键优化！）
                gc.collect()

                # V5.1: 清理临时上传的PDF文件
                if fp.startswith(tempfile.gettempdir()):
                    try:
                        os.unlink(fp)
                    except OSError:
                        pass

            status_text.empty()  # 清除状态文本
            st.success(f"✅ 审核完成！共审核 **{len(detail_results)}** 个文件")

            # 结果汇总表格
            df_summary = pd.DataFrame([
                {k: v for k, v in r.items() if not k.startswith("_")}
                for r in detail_results
            ])
            st.dataframe(df_summary, use_container_width=True, hide_index=True)

            # 统计概览
            col_s1, col_s2, col_s3 = st.columns(3)
            pass_count = sum(1 for r in detail_results if "合格" in r["总体结论"])
            fail_count = sum(1 for r in detail_results if "不合格" in r["总体结论"])
            warn_count = len(detail_results) - pass_count - fail_count

            with col_s1:
                st.metric("✅ 合格", pass_count, delta_color="normal")
            with col_s2:
                st.metric("⚠️ 需补充", warn_count, delta_color="off")
            with col_s3:
                st.metric("❌ 不合格", fail_count, delta_color="inverse")

            # 详细展开器
            st.subheader("📋 各文件详细审核报告")
            for idx, res in enumerate(detail_results):
                d = res["_detail"]
                expander_title = (
                    f"【{idx+1}】{res['文件名']} — "
                    f"类型:{res['物料类型']} | "
                    f"结论:{res['总体结论']} | "
                    f"问题数:{res['问题数量']}"
                )
                with st.expander(expander_title, expanded=(res["问题数量"] > 0)):
                    st.markdown(f"**文件类型:** {d['file_type_note']}")
                    st.markdown(f"**物料类型:** {d['mat_type_cn']} | **需电气性能测试:** {'是' if d['needs_elec'] else '否'}")

                    # PDF页面结构分析（V4.0新增）
                    st.subheader("📄 PDF页面结构分析（V4.0）")
                    if d.get("page_analysis"):
                        page_df = pd.DataFrame(d["page_analysis"])
                        st.dataframe(page_df, use_container_width=True, hide_index=True)
                    else:
                        st.text("⏱️ 无页面结构分析数据")

                    # 文件完整性详情
                    st.subheader("1️⃣ 文件完整性检验")
                    comp_data = d.get("completeness") or {}
                    comp_items = comp_data.get("items")
                    if comp_items:
                        comp_df = pd.DataFrame(comp_items)
                        st.dataframe(comp_df, hide_index=True, use_container_width=True)
                    else:
                        st.text(f"⏱ 无完整性数据 ({comp_data.get('status', '未知')})")

                    # RoHS详情
                    if rohs_check:
                        st.subheader("2️⃣ RoHS合规性检验")
                        rohs_sub = (d.get("rohs") or {}).get("sub_items", {})
                        if rohs_sub:
                            for k, v in rohs_sub.items():
                                st.text(f"{k}: {v}")
                        else:
                            st.text("⏱ 无RoHS数据")

                    # CPK详情
                    if cpk_check:
                        st.subheader("3️⃣ CPK合规性检验")
                        cpk_data = d.get("cpk") or {}
                        cpk_sub = cpk_data.get("sub_items", {})
                        if cpk_sub:
                            for k, v in cpk_sub.items():
                                st.text(f"{k}: {v}")
                        else:
                            st.text("⏱ 无CPK数据")
                        cpk_vals = cpk_data.get("cpk_values", [])
                        if cpk_vals:
                            st.text(f"检测到的CPK值: {cpk_vals}")

                    # 尺寸对应性详情
                    if dim_check:
                        st.subheader("4️⃣ 尺寸公差对应性检验")
                        dim_sub = (d.get("dimension") or {}).get("sub_items", {})
                        if dim_sub:
                            for k, v in dim_sub.items():
                                st.text(f"{k}: {v}")
                        else:
                            st.text("⏱ 无尺寸数据")

                    # 报告时效性详情
                    if validity_check:
                        st.subheader("5️⃣ 报告时效性检验")
                        val_sub = (d.get("validity") or {}).get("sub_items", {})
                        if val_sub:
                            for k, v in val_sub.items():
                                st.text(f"{k}: {v}")
                        else:
                            st.text("⏱ 无时效性数据")

                    # V6.2 新增：目录勾选状态
                    if d.get("catalog_check"):
                        cat = d["catalog_check"]
                        if not isinstance(cat, dict):
                            continue
                        st.subheader("6️⃣ 目录/Catalog勾选状态（V6.2）")
                        st.markdown(f"**判定结果:** {cat.get('status', 'N/A')}")
                        if cat.get('details'):
                            st.text(cat['details'])
                        if cat.get("has_catalog_page"):
                            col_cat1, col_cat2 = st.columns(2)
                            with col_cat1:
                                st.metric("已勾选", cat.get("checked_count", 0))
                            with col_cat2:
                                st.metric("未勾选", cat.get("unchecked_count", 0))
                        if cat.get("unchecked_items"):
                            st.text(f"未勾选项: {', '.join(cat['unchecked_items'][:10])}")

                    # V6.2 新增：料号&物料名称跨表一致性
                    if d.get("part_consistency"):
                        pc = d["part_consistency"]
                        if not isinstance(pc, dict):
                            continue
                        st.subheader("7️⃣ 料号&物料名称跨表一致性（V6.2）")
                        st.markdown(f"**整体判定:** {pc.get('overall_status', 'N/A')}")
                        # 显示封面参考信息
                        cover = pc.get("cover_info", {})
                        if cover.get("part_number") or cover.get("material_name"):
                            st.markdown(f"**封面参考信息:** 料号=`{cover.get('part_number', '未提取')}` | 名称=`{cover.get('material_name', '未提取')}`")
                        # 一致性检查表格
                        consistency_checks = pc.get("consistency_checks", [])
                        if consistency_checks:
                            pc_rows = []
                            for cc in consistency_checks:
                                pc_rows.append({
                                    "报告类型": cc["table_type"],
                                    "页码": cc["page_num"],
                                    "料号匹配": cc.get("pn_match", ""),
                                    "名称匹配": cc.get("name_match", ""),
                                })
                            pc_df = pd.DataFrame(pc_rows)
                            st.dataframe(pc_df, use_container_width=True, hide_index=True)

                    # V5.9.2 新增：封面供应商信息完整性
                    if d.get("supplier_check"):
                        sc = d["supplier_check"]
                        if isinstance(sc, dict):
                            st.subheader("7️⃣b 封面供应商信息完整性（V5.9.2）")
                            st.markdown(f"**整体判定:** {sc.get('overall_status', 'N/A')}")
                            if sc.get("supplier_name_found"):
                                st.markdown(f"**检测到供应商:** `{sc['supplier_name_found']}`")
                            # 逐项显示检查结果
                            sc_checks = sc.get("checks", [])
                            if sc_checks:
                                sc_rows = []
                                for c in sc_checks:
                                    sc_rows.append({
                                        "检查项": c.get("name", ""),
                                        "状态": c.get("status", ""),
                                        "详情": c.get("detail", "")[:80],
                                    })
                                sc_df = pd.DataFrame(sc_rows)
                                st.dataframe(sc_df, use_container_width=True, hide_index=True)

                    # 最终处理建议
                    st.subheader("8️⃣ 检验结论与处理建议")
                    st.markdown(f"**{d['final']['verdict']}**")
                    st.markdown(d['final']['suggestion'])

            # 下载Excel
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                excel_buf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows

                wb = Workbook()
                
                # --- Sheet1: 审核汇总 ---
                ws_summary = wb.active
                ws_summary.title = "审核汇总"
                for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        cell = ws_summary.cell(row=r_idx+1, column=c_idx+1, value=value)
                        if r_idx == 0:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")

                # --- Sheet2: 错误汇总（V5.3新增）---
                ws_error = wb.create_sheet("错误汇总")
                error_rows = []
                error_header_written = False
                for res in detail_results:
                    dd = res.get("_detail", {})
                    issues_found = []
                    
                    # 从各检查项中收集问题
                    checks = [
                        ("文件完整性", dd.get("completeness")),
                        ("RoHS合规", dd.get("rohs")),
                        ("CPK合规", dd.get("cpk")),
                        ("尺寸对应性", dd.get("dimension")),
                        ("报告时效性", dd.get("validity")),
                        ("目录勾选", dd.get("catalog_check")),
                        ("料号一致性", dd.get("part_consistency")),
                        ("螺丝图纸要求", dd.get("screw_check")),  # V5.9.0
                        ("供应商信息", dd.get("supplier_check")),  # V5.9.2
                    ]
                    
                    for check_name, check_data in checks:
                        if not check_data or not isinstance(check_data, dict):
                            continue
                        # 收集issues
                        for issue in check_data.get("issues", []):
                            issues_found.append({
                                "文件名": res.get("文件名", ""),
                                "检查类别": check_name,
                                "问题描述": str(issue)[:200],
                                "严重程度": "❌ 不合格" if "不合格" in str(issue) else "⚠️ 警告",
                                "建议操作": "请人工核实并补充相应资料或重新提交",
                            })
                        # sub_items中的失败项
                        for sub_k, sub_v in check_data.get("sub_items", {}).items():
                            if isinstance(sub_v, str) and ("❌" in sub_v or "不合格" in sub_v or "缺失" in sub_v):
                                issues_found.append({
                                    "文件名": res.get("文件名", ""),
                                    "检查类别": check_name,
                                    "问题描述": f"{sub_k}: {sub_v}"[:200],
                                    "严重程度": "❌ 不合格" if "❌" in sub_v else "⚠️ 警告",
                                    "建议操作": "请人工核实并补充相应资料或重新提交",
                                })
                        
                    # 最终判定中的问题
                    final = dd.get("final", {})
                    if "不合格" in final.get("verdict", "") or "异常" in final.get("verdict", ""):
                        issues_found.append({
                            "文件名": res.get("文件名", ""),
                            "检查类别": "最终结论",
                            "问题描述": final.get("verdict", "") + " | " + final.get("suggestion", "")[:150],
                            "严重程度": "❌ 关键问题" if "不合格" in final.get("verdict", "") else "⚠️ 异常",
                            "建议操作": final.get("suggestion", "请人工复核")[:200],
                        })
                    
                    error_rows.extend(issues_found)
                
                if error_rows:
                    # 写表头
                    err_headers = ["序号", "文件名", "检查类别", "问题描述", "严重程度", "建议操作"]
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    red_font = Font(color="9C0006")
                    bold_font = Font(bold=True)
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    for c_idx, h in enumerate(err_headers):
                        cell = ws_error.cell(row=1, column=c_idx+1, value=h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    for r_idx, erow in enumerate(error_rows):
                        for c_idx, key in enumerate(err_headers):
                            val = erow.get(key, "")
                            if c_idx == 0:
                                val = r_idx + 1
                            cell = ws_error.cell(row=r_idx+2, column=c_idx+1, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            # 根据严重程度着色
                            sev = str(erow.get("严重程度", ""))
                            if "关键" in sev or "不合格" in sev:
                                cell.fill = red_fill
                                cell.font = red_font
                            elif "警告" in sev or "⚠️" in sev:
                                cell.fill = yellow_fill
                    
                    # 设置列宽
                    ws_error.column_dimensions['A'].width = 6
                    ws_error.column_dimensions['B'].width = 30
                    ws_error.column_dimensions['C'].width = 14
                    ws_error.column_dimensions['D'].width = 50
                    ws_error.column_dimensions['E'].width = 12
                    ws_error.column_dimensions['F'].width = 35
                else:
                    ws_error.cell(row=1, column=1, value="✅ 所有文件均通过审核，无错误项！")
                    ws_error.cell(row=1, column=1).font = Font(bold=True, size=14, color="006400")
                
                wb.save(excel_buf.name)
                with open(excel_buf.name, "rb") as ef:
                    st.download_button(
                        label="📥 下载审核汇总 Excel（含错误汇总Sheet）",
                        data=ef,
                        file_name=f"封样审核汇总_V{version}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col_dl2:
                full_report = f"# 封样检验审核报告\n\n"
                full_report += f"**标准版本:** V{version}\n"
                full_report += f"**审核时间:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                full_report += f"**共审核:** {len(detail_results)} 个文件\n\n---\n\n"
                for i, res in enumerate(detail_results):
                    d = res["_detail"]
                    full_report += f"## 【{i+1}】{res['文件名']}\n\n"
                    full_report += f"**文件类型:** {d['file_type_note']}\n"
                    full_report += f"**物料类型:** {res['物料类型']}\n"
                    full_report += f"**总体结论:** {res['总体结论']}\n\n"
                    full_report += f"### 处理建议\n\n{d['final']['suggestion']}\n\n---\n\n"
                st.download_button(
                    label="📥 下载完整报告 Markdown",
                    data=full_report.encode("utf-8"),
                    file_name=f"封样审核报告_V{version}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                    mime="text/markdown"
                )

            # 清理临时文件
            for fp, _ in all_paths:
                if fp.startswith(tempfile.gettempdir()):
                    try:
                        os.unlink(fp)
                    except OSError:
                        pass
hide_style = "<style>#MainMenu {visibility: hidden;} footer {visibility: hidden;}</style>"
st.markdown(hide_style, unsafe_allow_html=True)
